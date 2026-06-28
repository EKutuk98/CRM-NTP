#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM SİSTEMİ — TIER 1 + TIER 2 + TIER 3
Müşteri Yönetimi + Dashboard + Raporlar & Export
"""

import sys
import sqlite3
import random
import warnings
from datetime import datetime, timedelta
from contextlib import contextmanager
import csv

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QDialog, QLabel,
    QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox, QMessageBox,
    QTabWidget, QFrame, QHeaderView, QTextEdit, QDateEdit, QFileDialog,
    QCalendarWidget
)
from PyQt5.QtCore import Qt, QTimer, QDate
from PyQt5.QtGui import QFont, QColor

import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

warnings.filterwarnings('ignore')

# ═════════════════════════════════════════════════════════════════════════════
# RENKLER
# ═════════════════════════════════════════════════════════════════════════════

COLORS = {
    'bg_main': '#0f0f1a',
    'bg_secondary': '#1a1a2e',
    'bg_tertiary': '#1e1e35',
    'primary': '#6366f1',
    'primary_light': '#818cf8',
    'secondary': '#a5b4fc',
    'success': '#10b981',
    'warning': '#f59e0b',
    'danger': '#ef4444',
    'text_main': '#ffffff',
    'text_sec': '#9ca3af',
    'border': '#2d2d44',
}

STYLESHEET = f"""
QMainWindow {{ background-color: {COLORS['bg_main']}; }}
QWidget {{ background-color: {COLORS['bg_main']}; color: {COLORS['text_main']}; }}

QPushButton {{
    background-color: {COLORS['primary']};
    color: white;
    border: none;
    padding: 10px 20px;
    border-radius: 10px;
    font-weight: bold;
    font-size: 12px;
}}
QPushButton:hover {{ background-color: {COLORS['primary_light']}; }}
QPushButton:pressed {{ background-color: #4f46e5; }}

QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {{
    background-color: {COLORS['bg_secondary']};
    border: 1px solid {COLORS['border']};
    border-radius: 8px;
    padding: 8px;
    color: {COLORS['text_main']};
    font-size: 12px;
}}
QLineEdit:focus {{ border: 1px solid {COLORS['primary']}; }}

QTableWidget {{
    background-color: {COLORS['bg_tertiary']};
    alternate-background-color: {COLORS['bg_secondary']};
    gridline-color: {COLORS['border']};
    border: 1px solid {COLORS['border']};
    border-radius: 8px;
}}
QTableWidget::item {{ padding: 8px; color: {COLORS['text_main']}; }}
QTableWidget::item:selected {{ background-color: {COLORS['primary']}; }}
QHeaderView::section {{
    background-color: {COLORS['bg_secondary']};
    color: {COLORS['secondary']};
    padding: 8px;
    border: none;
    font-weight: bold;
}}

QTabWidget::pane {{ border: 1px solid {COLORS['border']}; }}
QTabBar::tab {{
    background-color: {COLORS['bg_secondary']};
    color: {COLORS['text_sec']};
    padding: 8px 20px;
    border-radius: 6px;
    margin-right: 5px;
}}
QTabBar::tab:selected {{
    background-color: {COLORS['primary']};
    color: white;
}}

QLabel {{ color: {COLORS['text_main']}; }}
QDialog {{ background-color: {COLORS['bg_main']}; }}
"""

# ═════════════════════════════════════════════════════════════════════════════
# VERİTABANI
# ═════════════════════════════════════════════════════════════════════════════

class DatabaseManager:
    def __init__(self, db_name='crm_sistema.db'):
        self.db_name = db_name
        self.create_tables()
        self.load_sample_data()
    
    @contextmanager
    def get_connection(self):
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()
    
    def create_tables(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kullanicilar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kullanici_adi TEXT UNIQUE NOT NULL,
                    sifre TEXT NOT NULL,
                    ad TEXT NOT NULL,
                    soyad TEXT NOT NULL,
                    rol TEXT DEFAULT 'user'
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS musteriler (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ad TEXT NOT NULL,
                    email TEXT UNIQUE NOT NULL,
                    telefon TEXT,
                    sirket TEXT,
                    durum TEXT DEFAULT 'Aktif',
                    toplam_harcama REAL DEFAULT 0,
                    katilim_tarihi TEXT
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS satislar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER NOT NULL,
                    urun TEXT NOT NULL,
                    tutar REAL NOT NULL,
                    adet INTEGER NOT NULL,
                    toplam REAL NOT NULL,
                    durum TEXT DEFAULT 'Beklemede',
                    tarih TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS destekler (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER NOT NULL,
                    konu TEXT NOT NULL,
                    aciklama TEXT,
                    oncelik TEXT DEFAULT 'Orta',
                    durum TEXT DEFAULT 'Acik',
                    tarih TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS puanlar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER NOT NULL,
                    puan INTEGER DEFAULT 0,
                    seviye TEXT DEFAULT 'Bronz',
                    son_guncelleme TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS hediyeler (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER NOT NULL,
                    hediye_adi TEXT NOT NULL,
                    puan_maliyeti INTEGER NOT NULL,
                    durum TEXT DEFAULT 'Beklemede',
                    tarih TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS odul_katalog (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ad TEXT NOT NULL,
                    aciklama TEXT,
                    puan_maliyeti INTEGER NOT NULL,
                    stok INTEGER DEFAULT 100
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS email_sablonlari (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ad TEXT NOT NULL,
                    konu TEXT NOT NULL,
                    icerik TEXT NOT NULL,
                    kategori TEXT DEFAULT 'Genel'
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS email_gecmis (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER NOT NULL,
                    sablon_id INTEGER,
                    konu TEXT NOT NULL,
                    icerik TEXT,
                    durum TEXT DEFAULT 'Gonderildi',
                    tarih TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kampanyalar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ad TEXT NOT NULL,
                    aciklama TEXT,
                    tip TEXT DEFAULT 'Indirim',
                    indirim_orani REAL DEFAULT 0,
                    butce REAL DEFAULT 0,
                    hedef_segment TEXT DEFAULT 'Tum Musteriler',
                    basla_tarih TEXT,
                    bitis_tarih TEXT,
                    durum TEXT DEFAULT 'Taslak',
                    olusturma_tarihi TEXT
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kampanya_katilim (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kampanya_id INTEGER NOT NULL,
                    musteri_id INTEGER NOT NULL,
                    kullanildi INTEGER DEFAULT 0,
                    tarih TEXT,
                    FOREIGN KEY (kampanya_id) REFERENCES kampanyalar(id),
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS firsatlar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER NOT NULL,
                    baslik TEXT NOT NULL,
                    aciklama TEXT,
                    tahmini_deger REAL DEFAULT 0,
                    olasilik INTEGER DEFAULT 50,
                    asama TEXT DEFAULT 'Yeni',
                    son_iletisim TEXT,
                    beklenen_kapanis TEXT,
                    notlar TEXT,
                    olusturma_tarihi TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS etkinlikler (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    musteri_id INTEGER,
                    baslik TEXT NOT NULL,
                    aciklama TEXT,
                    tip TEXT DEFAULT 'Toplanti',
                    oncelik TEXT DEFAULT 'Orta',
                    tarih TEXT NOT NULL,
                    saat TEXT,
                    sure_dakika INTEGER DEFAULT 60,
                    konum TEXT,
                    durum TEXT DEFAULT 'Planlandi',
                    hatirlatici INTEGER DEFAULT 15,
                    olusturma_tarihi TEXT,
                    FOREIGN KEY (musteri_id) REFERENCES musteriler(id)
                )
            ''')
            
            try:
                cursor.execute("INSERT INTO kullanicilar (kullanici_adi, sifre, ad, soyad, rol) VALUES ('admin', 'admin123', 'Admin', 'User', 'admin')")
                cursor.execute("INSERT INTO kullanicilar (kullanici_adi, sifre, ad, soyad, rol) VALUES ('user', 'user123', 'Standard', 'User', 'user')")
            except sqlite3.IntegrityError:
                pass
    
    def load_sample_data(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # MUSTERILER
            cursor.execute('SELECT COUNT(*) FROM musteriler')
            musteri_var = cursor.fetchone()[0] > 0
            
            cursor.execute('SELECT id FROM musteriler ORDER BY id')
            mevcut_musteri_ids = [row[0] for row in cursor.fetchall()]
            
            if not musteri_var:
                musteriler = [
                    ("Ahmet Yilmaz", "ahmet@yilmaz.com", "+90 555 123 4567", "Yilmaz Holding", "Aktif"),
                    ("Ayse Demir", "ayse@demir.com", "+90 555 123 4568", "Demir Teknoloji", "Aktif"),
                    ("Mehmet Kaya", "mehmet@kaya.com", "+90 555 123 4569", "Kaya Grup", "Pasif"),
                    ("Zeynep Celik", "zeynep@celik.com", "+90 555 123 4570", "Celik Insaat", "Aktif"),
                    ("Can Ozturk", "can@ozturk.com", "+90 555 123 4571", "Ozturk Yazilim", "Aktif"),
                    ("Elif Sahin", "elif@sahin.com", "+90 555 123 4572", "Sahin Medya", "Beklemede"),
                    ("Fatih Kilic", "fatih@kilic.com", "+90 555 123 4573", "Kilic Imalat", "Aktif"),
                    ("Gul Arslan", "gul@arslan.com", "+90 555 123 4574", "Arslan Tekstil", "Aktif"),
                    ("Hakan Sen", "hakan@sen.com", "+90 555 123 4575", "Sen Danismanlik", "Pasif"),
                    ("Irem Turhan", "irem@turhan.com", "+90 555 123 4576", "Turhan Pazarlama", "Aktif"),
                    ("Jale Uysal", "jale@uysal.com", "+90 555 123 4577", "Uysal Turizm", "Aktif"),
                    ("Kerem Yaman", "kerem@yaman.com", "+90 555 123 4578", "Yaman Finans", "Aktif"),
                    ("Leyla Duman", "leyla@duman.com", "+90 555 123 4579", "Duman Tasarim", "Beklemede"),
                    ("Murat Gunes", "murat@gunes.com", "+90 555 123 4580", "Gunes Enerji", "Aktif"),
                    ("Nur Akcay", "nur@akcay.com", "+90 555 123 4581", "Akcay Proje", "Aktif"),
                    ("Omer Sarac", "omer@sarac.com", "+90 555 123 4582", "Sarac Insan Kaynaklari", "Aktif"),
                    ("Pinar Cetin", "pinar@cetin.com", "+90 555 123 4583", "Cetin Mimarlik", "Pasif"),
                    ("Recep Aydin", "recep@aydin.com", "+90 555 123 4584", "Aydin Peyzaj", "Aktif"),
                    ("Selin Kara", "selin@kara.com", "+90 555 123 4585", "Kara Hukuk", "Aktif"),
                    ("Tugba Alkan", "tugba@alkan.com", "+90 555 123 4586", "Alkan Gida", "Aktif"),
                ]
                
                for ad, email, tel, sirket, durum in musteriler:
                    cursor.execute('''
                        INSERT INTO musteriler (ad, email, telefon, sirket, durum, katilim_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (ad, email, tel, sirket, durum, (datetime.now() - timedelta(days=random.randint(0, 180))).strftime('%Y-%m-%d')))
                    mevcut_musteri_ids.append(cursor.lastrowid)
            
            # SATISLAR
            cursor.execute('SELECT COUNT(*) FROM satislar')
            if cursor.fetchone()[0] == 0 and mevcut_musteri_ids:
                urunler = ["Kurumsal Lisans", "Profesyonel Lisans", "Temel Lisans", "Premium Destek", "Bulut Depolama", "API Erisimi"]
                fiyatlar = [14999, 7499, 2999, 4499, 899, 2399]
                
                for _ in range(50):
                    musteri_id = random.choice(mevcut_musteri_ids)
                    urun = random.choice(urunler)
                    fiyat = fiyatlar[urunler.index(urun)]
                    adet = random.randint(1, 5)
                    durum_satis = random.choice(["Tamamlandi", "Isleniysor", "Beklemede"])
                    tarih = (datetime.now() - timedelta(days=random.randint(0, 90))).strftime('%Y-%m-%d')
                    
                    cursor.execute('''
                        INSERT INTO satislar (musteri_id, urun, tutar, adet, toplam, durum, tarih)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (musteri_id, urun, fiyat, adet, fiyat * adet, durum_satis, tarih))
            
            # DESTEKLER
            cursor.execute('SELECT COUNT(*) FROM destekler')
            if cursor.fetchone()[0] == 0 and mevcut_musteri_ids:
                konular = ["Giris sorunu", "Fatura sorgulama", "Teknik destek", "Yeni ozellik talebi", 
                           "Hata bildirimi", "Hesap yukseltme", "Sifre sifirla", "API dokumantasyon"]
                
                for _ in range(30):
                    musteri_id = random.choice(mevcut_musteri_ids)
                    konu = random.choice(konular)
                    oncelik = random.choice(["Dusuk", "Orta", "Yuksek", "Acil"])
                    durum_destek = random.choice(["Acik", "Islemde", "Cozuldu", "Kapali"])
                    tarih = (datetime.now() - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d')
                    
                    cursor.execute('''
                        INSERT INTO destekler (musteri_id, konu, aciklama, oncelik, durum, tarih)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (musteri_id, konu, f"Aciklama: {konu}", oncelik, durum_destek, tarih))
            
            # PUANLAR
            cursor.execute('SELECT COUNT(*) FROM puanlar')
            if cursor.fetchone()[0] == 0 and mevcut_musteri_ids:
                for musteri_id in mevcut_musteri_ids:
                    puan = random.randint(0, 5000)
                    if puan >= 3000:
                        seviye = "Platin"
                    elif puan >= 1500:
                        seviye = "Altin"
                    elif puan >= 500:
                        seviye = "Gumus"
                    else:
                        seviye = "Bronz"
                    
                    cursor.execute('''
                        INSERT INTO puanlar (musteri_id, puan, seviye, son_guncelleme)
                        VALUES (?, ?, ?, ?)
                    ''', (musteri_id, puan, seviye, datetime.now().strftime('%Y-%m-%d')))
            
            # ODUL KATALOG
            cursor.execute('SELECT COUNT(*) FROM odul_katalog')
            if cursor.fetchone()[0] == 0:
                oduller = [
                    ("Kahve Hediye Karti", "Starbucks 50 TRY", 500, 100),
                    ("Indirim Kuponu %10", "Sonraki alisverista", 300, 200),
                    ("Ucretsiz Kargo", "1 ay boyunca", 200, 500),
                    ("Premium Lisans 1 Ay", "Tum ozellikler acik", 2000, 50),
                    ("VIP Destek 6 Ay", "7/24 oncelikli destek", 3000, 20),
                    ("Tisort", "CRM Logolu", 800, 30),
                    ("Bilgisayar Cantasi", "Premium kalite", 5000, 10),
                    ("Powerbank", "20000 mAh", 1200, 40),
                    ("Bluetooth Kulaklik", "Kablosuz", 2500, 25),
                    ("Hediye Sepeti", "Cesitli urunler", 4000, 15),
                ]
                
                for ad, aciklama, puan, stok in oduller:
                    cursor.execute('''
                        INSERT INTO odul_katalog (ad, aciklama, puan_maliyeti, stok)
                        VALUES (?, ?, ?, ?)
                    ''', (ad, aciklama, puan, stok))
            
            # EMAIL SABLONLARI
            cursor.execute('SELECT COUNT(*) FROM email_sablonlari')
            if cursor.fetchone()[0] == 0:
                sablonlar = [
                    ("Hosgeldiniz", "Hosgeldiniz {ad}!", 
                     "Merhaba {ad},\n\nCRM sistemize hosgeldiniz! Sirketiniz {sirket} icin sizlere en iyi hizmeti sunmak icin buradayiz.\n\nSaygilarimizla,\nCRM Ekibi", "Hosgeldin"),
                    ("Tesekkur", "Tesekkur Ederiz {ad}", 
                     "Sayin {ad},\n\nSatin aldiginiz icin tesekkur ederiz. Memnuniyetiniz bizim icin onemli.\n\nSaygilarimizla,\nCRM Ekibi", "Satis"),
                    ("Pasif Musteri", "Sizi Ozledik {ad}", 
                     "Sayin {ad},\n\nBir suredir sizden haber alamiyoruz. Sizleri tekrar aramizda gormekten mutluluk duyariz.\n\nOzel indirim kuponunuz: HOSGELDIN10\n\nSaygilarimizla", "Pazarlama"),
                    ("Indirim Kampanyasi", "Size Ozel %20 Indirim!", 
                     "Sayin {ad},\n\nSizin icin ozel bir kampanyamiz var! 1 hafta gecerli %20 indirim firsatini kacirmayin.\n\nKupon Kodu: OZEL20\n\nSaygilarimizla", "Pazarlama"),
                    ("Yenileme Hatirlatma", "Aboneliginiz Bitiyor {ad}", 
                     "Sayin {ad},\n\n{sirket} icin aboneliginiz yakinda bitiyor. Hizmet kesintisi yasamamak icin yenileme yapmanizi oneririz.\n\nSaygilarimizla", "Hatirlatma"),
                    ("Puan Bildirim", "Puanlariniz Birikti {ad}!", 
                     "Sayin {ad},\n\nHesabinizda yeterli puan birikmis! Hediye katalogumuza goz atin.\n\nSaygilarimizla", "Bilgilendirme"),
                ]
                
                for ad, konu, icerik, kategori in sablonlar:
                    cursor.execute('''
                        INSERT INTO email_sablonlari (ad, konu, icerik, kategori)
                        VALUES (?, ?, ?, ?)
                    ''', (ad, konu, icerik, kategori))
            
            # KAMPANYALAR
            cursor.execute('SELECT COUNT(*) FROM kampanyalar')
            if cursor.fetchone()[0] == 0:
                bugun = datetime.now()
                kampanyalar = [
                    ("Yaz Indirim", "Yaz sezonu icin ozel indirim", "Indirim", 25.0, 50000, "Tum Musteriler", 
                     bugun.strftime('%Y-%m-%d'), (bugun + timedelta(days=30)).strftime('%Y-%m-%d'), "Aktif"),
                    ("Pasif Musteri Geri Kazanim", "Pasif musterileri geri getirme", "Indirim", 30.0, 25000, "Pasif Musteriler", 
                     bugun.strftime('%Y-%m-%d'), (bugun + timedelta(days=15)).strftime('%Y-%m-%d'), "Aktif"),
                    ("VIP Ozel Teklif", "Premium musteriler icin", "Premium", 15.0, 30000, "VIP Musteriler", 
                     bugun.strftime('%Y-%m-%d'), (bugun + timedelta(days=60)).strftime('%Y-%m-%d'), "Aktif"),
                    ("Yeni Yil Kampanyasi", "Yeni yil ozel firsat", "Indirim", 20.0, 75000, "Tum Musteriler", 
                     (bugun - timedelta(days=60)).strftime('%Y-%m-%d'), (bugun - timedelta(days=30)).strftime('%Y-%m-%d'), "Tamamlandi"),
                    ("Black Friday", "Black Friday ozel", "Indirim", 50.0, 100000, "Tum Musteriler", 
                     (bugun - timedelta(days=90)).strftime('%Y-%m-%d'), (bugun - timedelta(days=85)).strftime('%Y-%m-%d'), "Tamamlandi"),
                ]
                
                kampanya_ids = []
                for ad, aciklama, tip, indirim, butce, segment, basla, bitis, durum in kampanyalar:
                    cursor.execute('''
                        INSERT INTO kampanyalar (ad, aciklama, tip, indirim_orani, butce, hedef_segment, basla_tarih, bitis_tarih, durum, olusturma_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (ad, aciklama, tip, indirim, butce, segment, basla, bitis, durum, bugun.strftime('%Y-%m-%d')))
                    kampanya_ids.append(cursor.lastrowid)
                
                if mevcut_musteri_ids:
                    for kampanya_id in kampanya_ids:
                        katilim_sayi = random.randint(3, 15)
                        for _ in range(katilim_sayi):
                            musteri_id = random.choice(mevcut_musteri_ids)
                            kullanildi = random.choice([0, 1, 1])
                            cursor.execute('''
                                INSERT INTO kampanya_katilim (kampanya_id, musteri_id, kullanildi, tarih)
                                VALUES (?, ?, ?, ?)
                            ''', (kampanya_id, musteri_id, kullanildi, 
                                  (bugun - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d')))
            
            # FIRSATLAR
            cursor.execute('SELECT COUNT(*) FROM firsatlar')
            if cursor.fetchone()[0] == 0 and mevcut_musteri_ids:
                bugun = datetime.now()
                firsat_basliklar = [
                    "Kurumsal Lisans Yenileme", "Yeni Modul Satisi", "Premium Pakete Yukseltme",
                    "Cok Kullanicili Lisans", "API Entegrasyonu Projesi", "Egitim Hizmeti",
                    "Ozel Gelistirme Talebi", "Yillik Sozlesme Yenileme", "Bulut Gecisi Projesi",
                    "Danismanlik Hizmeti", "Mobil Uygulama Lisansi", "Birden Cok Lokasyon"
                ]
                
                asamalar = ["Yeni", "Iletisim", "Teklif", "Pazarlik", "Kazanildi", "Kaybedildi"]
                asama_olasilik = {"Yeni": 25, "Iletisim": 40, "Teklif": 60, "Pazarlik": 75, "Kazanildi": 100, "Kaybedildi": 0}
                
                for _ in range(25):
                    musteri_id = random.choice(mevcut_musteri_ids)
                    baslik = random.choice(firsat_basliklar)
                    tahmini = random.choice([5000, 10000, 15000, 25000, 50000, 75000, 100000])
                    asama = random.choice(asamalar)
                    olasilik = asama_olasilik[asama]
                    son_iletisim = (bugun - timedelta(days=random.randint(0, 30))).strftime('%Y-%m-%d')
                    beklenen = (bugun + timedelta(days=random.randint(7, 90))).strftime('%Y-%m-%d')
                    
                    cursor.execute('''
                        INSERT INTO firsatlar (musteri_id, baslik, aciklama, tahmini_deger, olasilik, asama, son_iletisim, beklenen_kapanis, notlar, olusturma_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (musteri_id, baslik, f"Detay: {baslik}", tahmini, olasilik, asama, 
                          son_iletisim, beklenen, f"Notlar: Musteri ile gorusme yapildi.", 
                          (bugun - timedelta(days=random.randint(0, 60))).strftime('%Y-%m-%d')))
            
            # ETKINLIKLER
            cursor.execute('SELECT COUNT(*) FROM etkinlikler')
            if cursor.fetchone()[0] == 0:
                bugun = datetime.now()
                etkinlik_basliklar = [
                    "Musteri Toplantisi", "Demo Sunumu", "Egitim Oturumu", "Strateji Toplantisi",
                    "Proje Kick-off", "Quarterly Review", "Sozlesme Imzalama", "Take Off Meeting",
                    "Teklif Sunumu", "Discovery Call", "Yenileme Gorusmesi", "Acil Toplanti"
                ]
                tipler = ["Toplanti", "Cagri", "Demo", "Egitim", "Sunum", "Etkinlik"]
                oncelikler = ["Dusuk", "Orta", "Yuksek", "Acil"]
                durumlar = ["Planlandi", "Tamamlandi", "Iptal", "Ertelendi"]
                
                for _ in range(20):
                    musteri_id = random.choice(mevcut_musteri_ids) if mevcut_musteri_ids and random.random() > 0.2 else None
                    baslik = random.choice(etkinlik_basliklar)
                    tip = random.choice(tipler)
                    oncelik = random.choice(oncelikler)
                    gun_ofset = random.randint(-15, 30)
                    tarih = (bugun + timedelta(days=gun_ofset)).strftime('%Y-%m-%d')
                    saat = f"{random.randint(9, 17):02d}:{random.choice(['00', '15', '30', '45'])}"
                    sure = random.choice([30, 60, 90, 120])
                    durum = random.choice(durumlar) if gun_ofset < 0 else "Planlandi"
                    
                    cursor.execute('''
                        INSERT INTO etkinlikler (musteri_id, baslik, aciklama, tip, oncelik, tarih, saat, sure_dakika, konum, durum, hatirlatici, olusturma_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (musteri_id, baslik, f"Aciklama: {baslik}", tip, oncelik, tarih, saat, sure, 
                          random.choice(["Ofis", "Online (Zoom)", "Musteri Lokasyonu", "Toplanti Salonu"]),
                          durum, random.choice([15, 30, 60]), bugun.strftime('%Y-%m-%d')))
            
    
    def musteri_ekle(self, ad, email, telefon, sirket, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO musteriler (ad, email, telefon, sirket, durum, katilim_tarihi)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (ad, email, telefon, sirket, durum, datetime.now().strftime('%Y-%m-%d')))
            return cursor.lastrowid
    
    def musterileri_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM musteriler ORDER BY id DESC')
            return [dict(row) for row in cursor.fetchall()]
    
    def musteri_guncelle(self, musteri_id, ad, email, telefon, sirket, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE musteriler SET ad=?, email=?, telefon=?, sirket=?, durum=?
                WHERE id=?
            ''', (ad, email, telefon, sirket, durum, musteri_id))
    
    def musteri_sil(self, musteri_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM satislar WHERE musteri_id=?', (musteri_id,))
            cursor.execute('DELETE FROM destekler WHERE musteri_id=?', (musteri_id,))
            cursor.execute('DELETE FROM musteriler WHERE id=?', (musteri_id,))
    
    def satis_ekle(self, musteri_id, urun, tutar, adet, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            toplam = tutar * adet
            cursor.execute('''
                INSERT INTO satislar (musteri_id, urun, tutar, adet, toplam, durum, tarih)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (musteri_id, urun, tutar, adet, toplam, durum, datetime.now().strftime('%Y-%m-%d')))
            return cursor.lastrowid
    
    def satislari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT s.*, m.ad as musteri_adi 
                FROM satislar s 
                JOIN musteriler m ON s.musteri_id = m.id
                ORDER BY s.id DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def satis_guncelle(self, satis_id, urun, tutar, adet, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            toplam = tutar * adet
            cursor.execute('''
                UPDATE satislar SET urun=?, tutar=?, adet=?, toplam=?, durum=?
                WHERE id=?
            ''', (urun, tutar, adet, toplam, durum, satis_id))
    
    def satis_sil(self, satis_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM satislar WHERE id=?', (satis_id,))
    
    def destek_ekle(self, musteri_id, konu, aciklama, oncelik, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO destekler (musteri_id, konu, aciklama, oncelik, durum, tarih)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (musteri_id, konu, aciklama, oncelik, durum, datetime.now().strftime('%Y-%m-%d')))
            return cursor.lastrowid
    
    def destekleri_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.*, m.ad as musteri_adi 
                FROM destekler d 
                JOIN musteriler m ON d.musteri_id = m.id
                ORDER BY d.id DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def destek_guncelle(self, destek_id, oncelik, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE destekler SET oncelik=?, durum=?
                WHERE id=?
            ''', (oncelik, durum, destek_id))
    
    def destek_sil(self, destek_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM destekler WHERE id=?', (destek_id,))
    
    # PUAN & HEDIYE METODLARI
    
    def puanlari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT p.*, m.ad as musteri_adi 
                FROM puanlar p 
                JOIN musteriler m ON p.musteri_id = m.id
                ORDER BY p.puan DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def musteri_puan_getir(self, musteri_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM puanlar WHERE musteri_id=?', (musteri_id,))
            row = cursor.fetchone()
            return dict(row) if row else None
    
    def puan_ekle(self, musteri_id, puan_miktar):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM puanlar WHERE musteri_id=?', (musteri_id,))
            mevcut = cursor.fetchone()
            
            if mevcut:
                yeni_puan = mevcut['puan'] + puan_miktar
                if yeni_puan >= 3000:
                    seviye = "Platin"
                elif yeni_puan >= 1500:
                    seviye = "Altin"
                elif yeni_puan >= 500:
                    seviye = "Gumus"
                else:
                    seviye = "Bronz"
                
                cursor.execute('''
                    UPDATE puanlar SET puan=?, seviye=?, son_guncelleme=?
                    WHERE musteri_id=?
                ''', (yeni_puan, seviye, datetime.now().strftime('%Y-%m-%d'), musteri_id))
            else:
                seviye = "Bronz" if puan_miktar < 500 else "Gumus"
                cursor.execute('''
                    INSERT INTO puanlar (musteri_id, puan, seviye, son_guncelleme)
                    VALUES (?, ?, ?, ?)
                ''', (musteri_id, puan_miktar, seviye, datetime.now().strftime('%Y-%m-%d')))
    
    def puan_dusur(self, musteri_id, puan_miktar):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM puanlar WHERE musteri_id=?', (musteri_id,))
            mevcut = cursor.fetchone()
            
            if not mevcut or mevcut['puan'] < puan_miktar:
                return False
            
            yeni_puan = mevcut['puan'] - puan_miktar
            if yeni_puan >= 3000:
                seviye = "Platin"
            elif yeni_puan >= 1500:
                seviye = "Altin"
            elif yeni_puan >= 500:
                seviye = "Gumus"
            else:
                seviye = "Bronz"
            
            cursor.execute('''
                UPDATE puanlar SET puan=?, seviye=?, son_guncelleme=?
                WHERE musteri_id=?
            ''', (yeni_puan, seviye, datetime.now().strftime('%Y-%m-%d'), musteri_id))
            return True
    
    def odul_katalog_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM odul_katalog ORDER BY puan_maliyeti ASC')
            return [dict(row) for row in cursor.fetchall()]
    
    def odul_ekle_katalog(self, ad, aciklama, puan, stok):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO odul_katalog (ad, aciklama, puan_maliyeti, stok)
                VALUES (?, ?, ?, ?)
            ''', (ad, aciklama, puan, stok))
    
    def odul_sil(self, odul_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM odul_katalog WHERE id=?', (odul_id,))
    
    def hediye_ver(self, musteri_id, odul_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM odul_katalog WHERE id=?', (odul_id,))
            odul = cursor.fetchone()
            
            if not odul:
                return False, "Odul bulunamadi"
            
            if odul['stok'] <= 0:
                return False, "Stokta yok"
            
            # Puan kontrol
            cursor.execute('SELECT * FROM puanlar WHERE musteri_id=?', (musteri_id,))
            puan_row = cursor.fetchone()
            
            if not puan_row or puan_row['puan'] < odul['puan_maliyeti']:
                return False, "Yeterli puan yok"
            
            # Puanı düş
            yeni_puan = puan_row['puan'] - odul['puan_maliyeti']
            if yeni_puan >= 3000:
                seviye = "Platin"
            elif yeni_puan >= 1500:
                seviye = "Altin"
            elif yeni_puan >= 500:
                seviye = "Gumus"
            else:
                seviye = "Bronz"
            
            cursor.execute('''
                UPDATE puanlar SET puan=?, seviye=?, son_guncelleme=?
                WHERE musteri_id=?
            ''', (yeni_puan, seviye, datetime.now().strftime('%Y-%m-%d'), musteri_id))
            
            # Stok düş
            cursor.execute('UPDATE odul_katalog SET stok=stok-1 WHERE id=?', (odul_id,))
            
            # Hediye kaydı
            cursor.execute('''
                INSERT INTO hediyeler (musteri_id, hediye_adi, puan_maliyeti, durum, tarih)
                VALUES (?, ?, ?, ?, ?)
            ''', (musteri_id, odul['ad'], odul['puan_maliyeti'], "Onaylandi", datetime.now().strftime('%Y-%m-%d')))
            
            return True, "Basarili"
    
    def hediyeleri_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT h.*, m.ad as musteri_adi 
                FROM hediyeler h 
                JOIN musteriler m ON h.musteri_id = m.id
                ORDER BY h.id DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    # EMAIL METODLARI
    
    def email_sablonlari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM email_sablonlari ORDER BY id DESC')
            return [dict(row) for row in cursor.fetchall()]
    
    def email_sablon_ekle(self, ad, konu, icerik, kategori):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO email_sablonlari (ad, konu, icerik, kategori)
                VALUES (?, ?, ?, ?)
            ''', (ad, konu, icerik, kategori))
    
    def email_sablon_sil(self, sablon_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM email_sablonlari WHERE id=?', (sablon_id,))
    
    def email_gonder_kayit(self, musteri_id, sablon_id, konu, icerik):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO email_gecmis (musteri_id, sablon_id, konu, icerik, durum, tarih)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (musteri_id, sablon_id, konu, icerik, "Gonderildi", datetime.now().strftime('%Y-%m-%d %H:%M')))
    
    def email_gecmis_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT e.*, m.ad as musteri_adi, m.email as musteri_email
                FROM email_gecmis e 
                JOIN musteriler m ON e.musteri_id = m.id
                ORDER BY e.id DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    # KAMPANYA METODLARI
    
    def kampanyalari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM kampanyalar ORDER BY id DESC')
            return [dict(row) for row in cursor.fetchall()]
    
    def kampanya_ekle(self, ad, aciklama, tip, indirim, butce, segment, basla, bitis, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO kampanyalar (ad, aciklama, tip, indirim_orani, butce, hedef_segment, basla_tarih, bitis_tarih, durum, olusturma_tarihi)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (ad, aciklama, tip, indirim, butce, segment, basla, bitis, durum, datetime.now().strftime('%Y-%m-%d')))
            return cursor.lastrowid
    
    def kampanya_guncelle(self, kampanya_id, ad, aciklama, tip, indirim, butce, segment, basla, bitis, durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE kampanyalar SET ad=?, aciklama=?, tip=?, indirim_orani=?, butce=?, hedef_segment=?, basla_tarih=?, bitis_tarih=?, durum=?
                WHERE id=?
            ''', (ad, aciklama, tip, indirim, butce, segment, basla, bitis, durum, kampanya_id))
    
    def kampanya_sil(self, kampanya_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM kampanya_katilim WHERE kampanya_id=?', (kampanya_id,))
            cursor.execute('DELETE FROM kampanyalar WHERE id=?', (kampanya_id,))
    
    def kampanya_katilimcilari_getir(self, kampanya_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT k.*, m.ad as musteri_adi, m.email as musteri_email
                FROM kampanya_katilim k
                JOIN musteriler m ON k.musteri_id = m.id
                WHERE k.kampanya_id=?
                ORDER BY k.id DESC
            ''', (kampanya_id,))
            return [dict(row) for row in cursor.fetchall()]
    
    def kampanya_katilim_ekle(self, kampanya_id, musteri_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO kampanya_katilim (kampanya_id, musteri_id, kullanildi, tarih)
                VALUES (?, ?, ?, ?)
            ''', (kampanya_id, musteri_id, 0, datetime.now().strftime('%Y-%m-%d')))
    
    def kampanya_istatistik(self, kampanya_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) as toplam FROM kampanya_katilim WHERE kampanya_id=?', (kampanya_id,))
            toplam = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) as kullanan FROM kampanya_katilim WHERE kampanya_id=? AND kullanildi=1', (kampanya_id,))
            kullanan = cursor.fetchone()[0]
            
            donusum = (kullanan / toplam * 100) if toplam > 0 else 0
            
            return {
                'toplam_katilim': toplam,
                'kullanan': kullanan,
                'donusum_orani': donusum
            }
    
    # FIRSAT METODLARI
    
    def firsatlari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT f.*, m.ad as musteri_adi, m.sirket as musteri_sirket
                FROM firsatlar f
                JOIN musteriler m ON f.musteri_id = m.id
                ORDER BY f.olusturma_tarihi DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def firsat_ekle(self, musteri_id, baslik, aciklama, deger, olasilik, asama, son_iletisim, beklenen, notlar):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO firsatlar (musteri_id, baslik, aciklama, tahmini_deger, olasilik, asama, son_iletisim, beklenen_kapanis, notlar, olusturma_tarihi)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (musteri_id, baslik, aciklama, deger, olasilik, asama, son_iletisim, beklenen, notlar, datetime.now().strftime('%Y-%m-%d')))
            return cursor.lastrowid
    
    def firsat_guncelle(self, firsat_id, baslik, aciklama, deger, olasilik, asama, son_iletisim, beklenen, notlar):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE firsatlar SET baslik=?, aciklama=?, tahmini_deger=?, olasilik=?, asama=?, son_iletisim=?, beklenen_kapanis=?, notlar=?
                WHERE id=?
            ''', (baslik, aciklama, deger, olasilik, asama, son_iletisim, beklenen, notlar, firsat_id))
    
    def firsat_asama_guncelle(self, firsat_id, yeni_asama):
        asama_olasilik = {"Yeni": 25, "Iletisim": 40, "Teklif": 60, "Pazarlik": 75, "Kazanildi": 100, "Kaybedildi": 0}
        olasilik = asama_olasilik.get(yeni_asama, 50)
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE firsatlar SET asama=?, olasilik=?, son_iletisim=?
                WHERE id=?
            ''', (yeni_asama, olasilik, datetime.now().strftime('%Y-%m-%d'), firsat_id))
    
    def firsat_sil(self, firsat_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM firsatlar WHERE id=?', (firsat_id,))
    
    # ETKINLIK METODLARI
    
    def etkinlikleri_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT e.*, m.ad as musteri_adi
                FROM etkinlikler e
                LEFT JOIN musteriler m ON e.musteri_id = m.id
                ORDER BY e.tarih ASC, e.saat ASC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def etkinlik_ekle(self, musteri_id, baslik, aciklama, tip, oncelik, tarih, saat, sure, konum, durum, hatirlatici):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO etkinlikler (musteri_id, baslik, aciklama, tip, oncelik, tarih, saat, sure_dakika, konum, durum, hatirlatici, olusturma_tarihi)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (musteri_id, baslik, aciklama, tip, oncelik, tarih, saat, sure, konum, durum, hatirlatici, datetime.now().strftime('%Y-%m-%d')))
            return cursor.lastrowid
    
    def etkinlik_guncelle(self, etkinlik_id, musteri_id, baslik, aciklama, tip, oncelik, tarih, saat, sure, konum, durum, hatirlatici):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE etkinlikler SET musteri_id=?, baslik=?, aciklama=?, tip=?, oncelik=?, tarih=?, saat=?, sure_dakika=?, konum=?, durum=?, hatirlatici=?
                WHERE id=?
            ''', (musteri_id, baslik, aciklama, tip, oncelik, tarih, saat, sure, konum, durum, hatirlatici, etkinlik_id))
    
    def etkinlik_sil(self, etkinlik_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM etkinlikler WHERE id=?', (etkinlik_id,))
    
    def etkinlik_durum_guncelle(self, etkinlik_id, yeni_durum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE etkinlikler SET durum=? WHERE id=?', (yeni_durum, etkinlik_id))

# ═════════════════════════════════════════════════════════════════════════════
# DİYALOGLAR
# ═════════════════════════════════════════════════════════════════════════════

class MusteriDialog(QDialog):
    def __init__(self, parent=None, musteri=None):
        super().__init__(parent)
        self.musteri = musteri
        self.setWindowTitle("Musteri Duzenle" if musteri else "Musteri Ekle")
        self.setModal(True)
        self.setGeometry(300, 300, 500, 450)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Musteri Bilgileri")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Ad Soyad *"))
        self.ad_input = QLineEdit()
        layout.addWidget(self.ad_input)
        
        layout.addWidget(QLabel("E-posta *"))
        self.email_input = QLineEdit()
        layout.addWidget(self.email_input)
        
        layout.addWidget(QLabel("Telefon"))
        self.tel_input = QLineEdit()
        layout.addWidget(self.tel_input)
        
        layout.addWidget(QLabel("Sirket"))
        self.sirket_input = QLineEdit()
        layout.addWidget(self.sirket_input)
        
        layout.addWidget(QLabel("Durum"))
        self.durum_combo = QComboBox()
        self.durum_combo.addItems(["Aktif", "Pasif", "Beklemede"])
        layout.addWidget(self.durum_combo)
        
        if self.musteri:
            self.ad_input.setText(self.musteri['ad'])
            self.email_input.setText(self.musteri['email'])
            self.tel_input.setText(self.musteri['telefon'] or "")
            self.sirket_input.setText(self.musteri['sirket'] or "")
            self.durum_combo.setCurrentText(self.musteri['durum'])
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.ad_input.text() or not self.email_input.text():
            QMessageBox.warning(self, "Uyari", "Ad Soyad ve E-posta zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "ad": self.ad_input.text(),
            "email": self.email_input.text(),
            "telefon": self.tel_input.text(),
            "sirket": self.sirket_input.text(),
            "durum": self.durum_combo.currentText()
        }

class SatisDialog(QDialog):
    def __init__(self, parent=None, satis=None):
        super().__init__(parent)
        self.satis = satis
        self.setWindowTitle("Satis Duzenle" if satis else "Satis Ekle")
        self.setModal(True)
        self.setGeometry(300, 300, 450, 350)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Satis Bilgileri")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Urun *"))
        self.urun_input = QLineEdit()
        layout.addWidget(self.urun_input)
        
        layout.addWidget(QLabel("Tutar (TRY) *"))
        self.tutar_input = QDoubleSpinBox()
        self.tutar_input.setMaximum(1000000)
        layout.addWidget(self.tutar_input)
        
        layout.addWidget(QLabel("Adet *"))
        self.adet_input = QSpinBox()
        self.adet_input.setMinimum(1)
        layout.addWidget(self.adet_input)
        
        layout.addWidget(QLabel("Durum"))
        self.durum_combo = QComboBox()
        self.durum_combo.addItems(["Beklemede", "Isleniysor", "Tamamlandi"])
        layout.addWidget(self.durum_combo)
        
        if self.satis:
            self.urun_input.setText(self.satis['urun'])
            self.tutar_input.setValue(self.satis['tutar'])
            self.adet_input.setValue(self.satis['adet'])
            self.durum_combo.setCurrentText(self.satis['durum'])
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.urun_input.text() or self.tutar_input.value() <= 0:
            QMessageBox.warning(self, "Uyari", "Urun ve tutar zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "urun": self.urun_input.text(),
            "tutar": self.tutar_input.value(),
            "adet": self.adet_input.value(),
            "durum": self.durum_combo.currentText()
        }

class DestekDialog(QDialog):
    def __init__(self, parent=None, destek=None):
        super().__init__(parent)
        self.destek = destek
        self.setWindowTitle("Destek Duzenle" if destek else "Destek Ekle")
        self.setModal(True)
        self.setGeometry(300, 300, 500, 450)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Destek Talebi")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Konu *"))
        self.konu_input = QLineEdit()
        layout.addWidget(self.konu_input)
        
        layout.addWidget(QLabel("Aciklama"))
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(100)
        layout.addWidget(self.aciklama_input)
        
        layout.addWidget(QLabel("Oncelik"))
        self.oncelik_combo = QComboBox()
        self.oncelik_combo.addItems(["Dusuk", "Orta", "Yuksek", "Acil"])
        layout.addWidget(self.oncelik_combo)
        
        layout.addWidget(QLabel("Durum"))
        self.durum_combo = QComboBox()
        self.durum_combo.addItems(["Acik", "Islemde", "Cozuldu", "Kapali"])
        layout.addWidget(self.durum_combo)
        
        if self.destek:
            self.konu_input.setText(self.destek['konu'])
            self.aciklama_input.setText(self.destek['aciklama'] or "")
            self.oncelik_combo.setCurrentText(self.destek['oncelik'])
            self.durum_combo.setCurrentText(self.destek['durum'])
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.konu_input.text():
            QMessageBox.warning(self, "Uyari", "Konu zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "konu": self.konu_input.text(),
            "aciklama": self.aciklama_input.toPlainText(),
            "oncelik": self.oncelik_combo.currentText(),
            "durum": self.durum_combo.currentText()
        }

class OdulKatalogDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Odul Ekle")
        self.setModal(True)
        self.setGeometry(300, 300, 450, 400)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Yeni Odul Ekle")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Odul Adi *"))
        self.ad_input = QLineEdit()
        layout.addWidget(self.ad_input)
        
        layout.addWidget(QLabel("Aciklama"))
        self.aciklama_input = QLineEdit()
        layout.addWidget(self.aciklama_input)
        
        layout.addWidget(QLabel("Puan Maliyeti *"))
        self.puan_input = QSpinBox()
        self.puan_input.setMaximum(100000)
        self.puan_input.setMinimum(1)
        layout.addWidget(self.puan_input)
        
        layout.addWidget(QLabel("Stok *"))
        self.stok_input = QSpinBox()
        self.stok_input.setMaximum(10000)
        self.stok_input.setMinimum(1)
        layout.addWidget(self.stok_input)
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.ad_input.text():
            QMessageBox.warning(self, "Uyari", "Odul adi zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "ad": self.ad_input.text(),
            "aciklama": self.aciklama_input.text(),
            "puan": self.puan_input.value(),
            "stok": self.stok_input.value()
        }

class PuanEkleDialog(QDialog):
    def __init__(self, parent=None, musteri_adi=""):
        super().__init__(parent)
        self.setWindowTitle("Puan Ekle/Cikar")
        self.setModal(True)
        self.setGeometry(300, 300, 400, 250)
        self.setStyleSheet(STYLESHEET)
        self.musteri_adi = musteri_adi
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel(f"Musteri: {self.musteri_adi}")
        title.setFont(QFont("Arial", 12, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Islem"))
        self.islem_combo = QComboBox()
        self.islem_combo.addItems(["Puan Ekle", "Puan Cikar"])
        layout.addWidget(self.islem_combo)
        
        layout.addWidget(QLabel("Miktar *"))
        self.miktar_input = QSpinBox()
        self.miktar_input.setMaximum(100000)
        self.miktar_input.setMinimum(1)
        layout.addWidget(self.miktar_input)
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Uygula")
        kaydet_btn.clicked.connect(self.accept)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def veri_al(self):
        return {
            "islem": self.islem_combo.currentText(),
            "miktar": self.miktar_input.value()
        }

class HediyeVerDialog(QDialog):
    def __init__(self, parent=None, musteriler=None, oduller=None):
        super().__init__(parent)
        self.musteriler = musteriler or []
        self.oduller = oduller or []
        self.setWindowTitle("Hediye Ver")
        self.setModal(True)
        self.setGeometry(300, 300, 450, 300)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Musteriye Hediye Ver")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Musteri Sec *"))
        self.musteri_combo = QComboBox()
        for m in self.musteriler:
            self.musteri_combo.addItem(f"{m['ad']} (ID: {m['id']})", m['id'])
        layout.addWidget(self.musteri_combo)
        
        layout.addWidget(QLabel("Odul Sec *"))
        self.odul_combo = QComboBox()
        for o in self.oduller:
            if o['stok'] > 0:
                self.odul_combo.addItem(f"{o['ad']} - {o['puan_maliyeti']} puan (Stok: {o['stok']})", o['id'])
        layout.addWidget(self.odul_combo)
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Hediye Ver")
        kaydet_btn.clicked.connect(self.accept)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def veri_al(self):
        return {
            "musteri_id": self.musteri_combo.currentData(),
            "odul_id": self.odul_combo.currentData()
        }

class EmailSablonDialog(QDialog):
    def __init__(self, parent=None, sablon=None):
        super().__init__(parent)
        self.sablon = sablon
        self.setWindowTitle("Sablon Duzenle" if sablon else "Yeni Sablon")
        self.setModal(True)
        self.setGeometry(300, 300, 600, 500)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Email Sablonu")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Sablon Adi *"))
        self.ad_input = QLineEdit()
        layout.addWidget(self.ad_input)
        
        layout.addWidget(QLabel("Kategori"))
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems(["Genel", "Hosgeldin", "Satis", "Pazarlama", "Hatirlatma", "Bilgilendirme"])
        layout.addWidget(self.kategori_combo)
        
        layout.addWidget(QLabel("Konu *"))
        self.konu_input = QLineEdit()
        layout.addWidget(self.konu_input)
        
        layout.addWidget(QLabel("Icerik * (Kullanilabilir: {ad}, {email}, {sirket})"))
        self.icerik_input = QTextEdit()
        self.icerik_input.setMinimumHeight(200)
        layout.addWidget(self.icerik_input)
        
        if self.sablon:
            self.ad_input.setText(self.sablon['ad'])
            self.kategori_combo.setCurrentText(self.sablon['kategori'])
            self.konu_input.setText(self.sablon['konu'])
            self.icerik_input.setText(self.sablon['icerik'])
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.ad_input.text() or not self.konu_input.text() or not self.icerik_input.toPlainText():
            QMessageBox.warning(self, "Uyari", "Tum alanlar zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "ad": self.ad_input.text(),
            "kategori": self.kategori_combo.currentText(),
            "konu": self.konu_input.text(),
            "icerik": self.icerik_input.toPlainText()
        }

class EmailGonderDialog(QDialog):
    def __init__(self, parent=None, musteriler=None, sablonlar=None):
        super().__init__(parent)
        self.musteriler = musteriler or []
        self.sablonlar = sablonlar or []
        self.setWindowTitle("Email Gonder")
        self.setModal(True)
        self.setGeometry(300, 300, 700, 600)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Email Gonder")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Alici Tipi"))
        self.alici_tip_combo = QComboBox()
        self.alici_tip_combo.addItems(["Tek Musteri", "Tum Aktifler", "Tum Pasifler", "Tum Musteriler"])
        self.alici_tip_combo.currentTextChanged.connect(self.alici_tip_degisti)
        layout.addWidget(self.alici_tip_combo)
        
        self.musteri_label = QLabel("Musteri Sec *")
        layout.addWidget(self.musteri_label)
        self.musteri_combo = QComboBox()
        for m in self.musteriler:
            self.musteri_combo.addItem(f"{m['ad']} ({m['email']})", m['id'])
        layout.addWidget(self.musteri_combo)
        
        layout.addWidget(QLabel("Sablon Sec"))
        self.sablon_combo = QComboBox()
        self.sablon_combo.addItem("--- Manuel ---", None)
        for s in self.sablonlar:
            self.sablon_combo.addItem(f"[{s['kategori']}] {s['ad']}", s['id'])
        self.sablon_combo.currentIndexChanged.connect(self.sablon_secildi)
        layout.addWidget(self.sablon_combo)
        
        layout.addWidget(QLabel("Konu *"))
        self.konu_input = QLineEdit()
        layout.addWidget(self.konu_input)
        
        layout.addWidget(QLabel("Icerik *"))
        self.icerik_input = QTextEdit()
        self.icerik_input.setMinimumHeight(200)
        layout.addWidget(self.icerik_input)
        
        btn_layout = QHBoxLayout()
        gonder_btn = QPushButton("Gonder")
        gonder_btn.clicked.connect(self.gonder)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(gonder_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def alici_tip_degisti(self):
        tek = self.alici_tip_combo.currentText() == "Tek Musteri"
        self.musteri_label.setVisible(tek)
        self.musteri_combo.setVisible(tek)
    
    def sablon_secildi(self):
        sablon_id = self.sablon_combo.currentData()
        if sablon_id is None:
            return
        
        for s in self.sablonlar:
            if s['id'] == sablon_id:
                self.konu_input.setText(s['konu'])
                self.icerik_input.setText(s['icerik'])
                break
    
    def gonder(self):
        if not self.konu_input.text() or not self.icerik_input.toPlainText():
            QMessageBox.warning(self, "Uyari", "Konu ve Icerik zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "alici_tip": self.alici_tip_combo.currentText(),
            "musteri_id": self.musteri_combo.currentData(),
            "sablon_id": self.sablon_combo.currentData(),
            "konu": self.konu_input.text(),
            "icerik": self.icerik_input.toPlainText()
        }

class KampanyaDialog(QDialog):
    def __init__(self, parent=None, kampanya=None):
        super().__init__(parent)
        self.kampanya = kampanya
        self.setWindowTitle("Kampanya Duzenle" if kampanya else "Yeni Kampanya")
        self.setModal(True)
        self.setGeometry(300, 300, 550, 650)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Kampanya Bilgileri")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Kampanya Adi *"))
        self.ad_input = QLineEdit()
        layout.addWidget(self.ad_input)
        
        layout.addWidget(QLabel("Aciklama"))
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        layout.addWidget(self.aciklama_input)
        
        layout.addWidget(QLabel("Tip"))
        self.tip_combo = QComboBox()
        self.tip_combo.addItems(["Indirim", "Premium", "Hediye", "Sadakat", "Yeni Musteri"])
        layout.addWidget(self.tip_combo)
        
        layout.addWidget(QLabel("Indirim Orani (%)"))
        self.indirim_input = QDoubleSpinBox()
        self.indirim_input.setMaximum(100)
        self.indirim_input.setMinimum(0)
        layout.addWidget(self.indirim_input)
        
        layout.addWidget(QLabel("Butce (TRY)"))
        self.butce_input = QDoubleSpinBox()
        self.butce_input.setMaximum(10000000)
        self.butce_input.setMinimum(0)
        layout.addWidget(self.butce_input)
        
        layout.addWidget(QLabel("Hedef Segment"))
        self.segment_combo = QComboBox()
        self.segment_combo.addItems(["Tum Musteriler", "Aktif Musteriler", "Pasif Musteriler", "VIP Musteriler", "Yeni Musteriler"])
        layout.addWidget(self.segment_combo)
        
        layout.addWidget(QLabel("Baslangic Tarihi"))
        self.basla_date = QDateEdit()
        self.basla_date.setDate(QDate.currentDate())
        layout.addWidget(self.basla_date)
        
        layout.addWidget(QLabel("Bitis Tarihi"))
        self.bitis_date = QDateEdit()
        self.bitis_date.setDate(QDate.currentDate().addMonths(1))
        layout.addWidget(self.bitis_date)
        
        layout.addWidget(QLabel("Durum"))
        self.durum_combo = QComboBox()
        self.durum_combo.addItems(["Taslak", "Aktif", "Duraklatildi", "Tamamlandi", "Iptal"])
        layout.addWidget(self.durum_combo)
        
        if self.kampanya:
            self.ad_input.setText(self.kampanya['ad'])
            self.aciklama_input.setText(self.kampanya['aciklama'] or "")
            self.tip_combo.setCurrentText(self.kampanya['tip'])
            self.indirim_input.setValue(self.kampanya['indirim_orani'])
            self.butce_input.setValue(self.kampanya['butce'])
            self.segment_combo.setCurrentText(self.kampanya['hedef_segment'])
            
            try:
                basla_date = QDate.fromString(self.kampanya['basla_tarih'], "yyyy-MM-dd")
                bitis_date = QDate.fromString(self.kampanya['bitis_tarih'], "yyyy-MM-dd")
                self.basla_date.setDate(basla_date)
                self.bitis_date.setDate(bitis_date)
            except:
                pass
            
            self.durum_combo.setCurrentText(self.kampanya['durum'])
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.ad_input.text():
            QMessageBox.warning(self, "Uyari", "Kampanya adi zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "ad": self.ad_input.text(),
            "aciklama": self.aciklama_input.toPlainText(),
            "tip": self.tip_combo.currentText(),
            "indirim": self.indirim_input.value(),
            "butce": self.butce_input.value(),
            "segment": self.segment_combo.currentText(),
            "basla": self.basla_date.date().toString("yyyy-MM-dd"),
            "bitis": self.bitis_date.date().toString("yyyy-MM-dd"),
            "durum": self.durum_combo.currentText()
        }

class FirsatDialog(QDialog):
    def __init__(self, parent=None, musteriler=None, firsat=None):
        super().__init__(parent)
        self.musteriler = musteriler or []
        self.firsat = firsat
        self.setWindowTitle("Firsat Duzenle" if firsat else "Yeni Firsat")
        self.setModal(True)
        self.setGeometry(300, 300, 550, 700)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Firsat Bilgileri")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Musteri *"))
        self.musteri_combo = QComboBox()
        for m in self.musteriler:
            self.musteri_combo.addItem(f"{m['ad']} - {m['sirket'] or ''}", m['id'])
        layout.addWidget(self.musteri_combo)
        
        layout.addWidget(QLabel("Baslik *"))
        self.baslik_input = QLineEdit()
        layout.addWidget(self.baslik_input)
        
        layout.addWidget(QLabel("Aciklama"))
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        layout.addWidget(self.aciklama_input)
        
        layout.addWidget(QLabel("Tahmini Deger (TRY) *"))
        self.deger_input = QDoubleSpinBox()
        self.deger_input.setMaximum(10000000)
        self.deger_input.setMinimum(0)
        layout.addWidget(self.deger_input)
        
        layout.addWidget(QLabel("Olasilik (%)"))
        self.olasilik_input = QSpinBox()
        self.olasilik_input.setMaximum(100)
        self.olasilik_input.setMinimum(0)
        self.olasilik_input.setValue(50)
        layout.addWidget(self.olasilik_input)
        
        layout.addWidget(QLabel("Asama"))
        self.asama_combo = QComboBox()
        self.asama_combo.addItems(["Yeni", "Iletisim", "Teklif", "Pazarlik", "Kazanildi", "Kaybedildi"])
        layout.addWidget(self.asama_combo)
        
        layout.addWidget(QLabel("Son Iletisim Tarihi"))
        self.son_date = QDateEdit()
        self.son_date.setDate(QDate.currentDate())
        layout.addWidget(self.son_date)
        
        layout.addWidget(QLabel("Beklenen Kapanis Tarihi"))
        self.beklenen_date = QDateEdit()
        self.beklenen_date.setDate(QDate.currentDate().addDays(30))
        layout.addWidget(self.beklenen_date)
        
        layout.addWidget(QLabel("Notlar"))
        self.notlar_input = QTextEdit()
        self.notlar_input.setMaximumHeight(80)
        layout.addWidget(self.notlar_input)
        
        if self.firsat:
            # Müşteri seç
            for i in range(self.musteri_combo.count()):
                if self.musteri_combo.itemData(i) == self.firsat['musteri_id']:
                    self.musteri_combo.setCurrentIndex(i)
                    break
            
            self.baslik_input.setText(self.firsat['baslik'])
            self.aciklama_input.setText(self.firsat['aciklama'] or "")
            self.deger_input.setValue(self.firsat['tahmini_deger'])
            self.olasilik_input.setValue(self.firsat['olasilik'])
            self.asama_combo.setCurrentText(self.firsat['asama'])
            
            try:
                if self.firsat['son_iletisim']:
                    self.son_date.setDate(QDate.fromString(self.firsat['son_iletisim'], "yyyy-MM-dd"))
                if self.firsat['beklenen_kapanis']:
                    self.beklenen_date.setDate(QDate.fromString(self.firsat['beklenen_kapanis'], "yyyy-MM-dd"))
            except:
                pass
            
            self.notlar_input.setText(self.firsat['notlar'] or "")
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.baslik_input.text():
            QMessageBox.warning(self, "Uyari", "Baslik zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "musteri_id": self.musteri_combo.currentData(),
            "baslik": self.baslik_input.text(),
            "aciklama": self.aciklama_input.toPlainText(),
            "deger": self.deger_input.value(),
            "olasilik": self.olasilik_input.value(),
            "asama": self.asama_combo.currentText(),
            "son_iletisim": self.son_date.date().toString("yyyy-MM-dd"),
            "beklenen": self.beklenen_date.date().toString("yyyy-MM-dd"),
            "notlar": self.notlar_input.toPlainText()
        }

class EtkinlikDialog(QDialog):
    def __init__(self, parent=None, musteriler=None, etkinlik=None):
        super().__init__(parent)
        self.musteriler = musteriler or []
        self.etkinlik = etkinlik
        self.setWindowTitle("Etkinlik Duzenle" if etkinlik else "Yeni Etkinlik")
        self.setModal(True)
        self.setGeometry(300, 300, 550, 700)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        title = QLabel("Etkinlik Bilgileri")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Baslik *"))
        self.baslik_input = QLineEdit()
        layout.addWidget(self.baslik_input)
        
        layout.addWidget(QLabel("Aciklama"))
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        layout.addWidget(self.aciklama_input)
        
        layout.addWidget(QLabel("Musteri (Opsiyonel)"))
        self.musteri_combo = QComboBox()
        self.musteri_combo.addItem("--- Yok ---", None)
        for m in self.musteriler:
            self.musteri_combo.addItem(f"{m['ad']} - {m['sirket'] or ''}", m['id'])
        layout.addWidget(self.musteri_combo)
        
        layout.addWidget(QLabel("Tip"))
        self.tip_combo = QComboBox()
        self.tip_combo.addItems(["Toplanti", "Cagri", "Demo", "Egitim", "Sunum", "Etkinlik"])
        layout.addWidget(self.tip_combo)
        
        layout.addWidget(QLabel("Oncelik"))
        self.oncelik_combo = QComboBox()
        self.oncelik_combo.addItems(["Dusuk", "Orta", "Yuksek", "Acil"])
        self.oncelik_combo.setCurrentText("Orta")
        layout.addWidget(self.oncelik_combo)
        
        tarih_saat_layout = QHBoxLayout()
        tarih_widget = QWidget()
        tarih_layout = QVBoxLayout()
        tarih_layout.setContentsMargins(0, 0, 0, 0)
        tarih_layout.addWidget(QLabel("Tarih *"))
        self.tarih_date = QDateEdit()
        self.tarih_date.setDate(QDate.currentDate())
        tarih_layout.addWidget(self.tarih_date)
        tarih_widget.setLayout(tarih_layout)
        
        saat_widget = QWidget()
        saat_layout = QVBoxLayout()
        saat_layout.setContentsMargins(0, 0, 0, 0)
        saat_layout.addWidget(QLabel("Saat (HH:MM)"))
        self.saat_input = QLineEdit()
        self.saat_input.setText("09:00")
        saat_layout.addWidget(self.saat_input)
        saat_widget.setLayout(saat_layout)
        
        tarih_saat_layout.addWidget(tarih_widget)
        tarih_saat_layout.addWidget(saat_widget)
        layout.addLayout(tarih_saat_layout)
        
        layout.addWidget(QLabel("Sure (Dakika)"))
        self.sure_input = QSpinBox()
        self.sure_input.setMaximum(1440)
        self.sure_input.setMinimum(15)
        self.sure_input.setValue(60)
        self.sure_input.setSingleStep(15)
        layout.addWidget(self.sure_input)
        
        layout.addWidget(QLabel("Konum"))
        self.konum_input = QLineEdit()
        layout.addWidget(self.konum_input)
        
        layout.addWidget(QLabel("Durum"))
        self.durum_combo = QComboBox()
        self.durum_combo.addItems(["Planlandi", "Tamamlandi", "Iptal", "Ertelendi"])
        layout.addWidget(self.durum_combo)
        
        layout.addWidget(QLabel("Hatirlatici (Dakika once)"))
        self.hatirlatici_combo = QComboBox()
        self.hatirlatici_combo.addItems(["5", "15", "30", "60", "120", "1440"])
        self.hatirlatici_combo.setCurrentText("15")
        layout.addWidget(self.hatirlatici_combo)
        
        if self.etkinlik:
            self.baslik_input.setText(self.etkinlik['baslik'])
            self.aciklama_input.setText(self.etkinlik['aciklama'] or "")
            
            if self.etkinlik['musteri_id']:
                for i in range(self.musteri_combo.count()):
                    if self.musteri_combo.itemData(i) == self.etkinlik['musteri_id']:
                        self.musteri_combo.setCurrentIndex(i)
                        break
            
            self.tip_combo.setCurrentText(self.etkinlik['tip'])
            self.oncelik_combo.setCurrentText(self.etkinlik['oncelik'])
            
            try:
                if self.etkinlik['tarih']:
                    self.tarih_date.setDate(QDate.fromString(self.etkinlik['tarih'], "yyyy-MM-dd"))
            except:
                pass
            
            self.saat_input.setText(self.etkinlik['saat'] or "09:00")
            self.sure_input.setValue(self.etkinlik['sure_dakika'] or 60)
            self.konum_input.setText(self.etkinlik['konum'] or "")
            self.durum_combo.setCurrentText(self.etkinlik['durum'])
            self.hatirlatici_combo.setCurrentText(str(self.etkinlik['hatirlatici']))
        
        btn_layout = QHBoxLayout()
        kaydet_btn = QPushButton("Kaydet")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def kaydet(self):
        if not self.baslik_input.text():
            QMessageBox.warning(self, "Uyari", "Baslik zorunludur!")
            return
        self.accept()
    
    def veri_al(self):
        return {
            "musteri_id": self.musteri_combo.currentData(),
            "baslik": self.baslik_input.text(),
            "aciklama": self.aciklama_input.toPlainText(),
            "tip": self.tip_combo.currentText(),
            "oncelik": self.oncelik_combo.currentText(),
            "tarih": self.tarih_date.date().toString("yyyy-MM-dd"),
            "saat": self.saat_input.text(),
            "sure": self.sure_input.value(),
            "konum": self.konum_input.text(),
            "durum": self.durum_combo.currentText(),
            "hatirlatici": int(self.hatirlatici_combo.currentText())
        }

class RaporDialog(QDialog):
    def __init__(self, parent=None, rapor_tipi="musteri"):
        super().__init__(parent)
        self.rapor_tipi = rapor_tipi
        self.setWindowTitle(f"{rapor_tipi.capitalize()} Raporu")
        self.setModal(True)
        self.setGeometry(300, 300, 500, 250)
        self.setStyleSheet(STYLESHEET)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        
        layout.addWidget(QLabel("Baslangic Tarihi"))
        self.basla_date = QDateEdit()
        self.basla_date.setDate(QDate.currentDate().addMonths(-1))
        layout.addWidget(self.basla_date)
        
        layout.addWidget(QLabel("Bitis Tarihi"))
        self.bitis_date = QDateEdit()
        self.bitis_date.setDate(QDate.currentDate())
        layout.addWidget(self.bitis_date)
        
        layout.addWidget(QLabel("Format"))
        self.format_combo = QComboBox()
        self.format_combo.addItems(["CSV", "Excel"])
        layout.addWidget(self.format_combo)
        
        layout.addStretch()
        
        btn_layout = QHBoxLayout()
        indir_btn = QPushButton("Indir")
        indir_btn.clicked.connect(self.indir)
        iptal_btn = QPushButton("Iptal")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(indir_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def indir(self):
        self.accept()
    
    def veri_al(self):
        return {
            "basla": self.basla_date.date().toString("yyyy-MM-dd"),
            "bitis": self.bitis_date.date().toString("yyyy-MM-dd"),
            "format": self.format_combo.currentText()
        }

# ═════════════════════════════════════════════════════════════════════════════
# ANA PENCERE
# ═════════════════════════════════════════════════════════════════════════════

class CRMMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.kullanici = None
        
        self.setWindowTitle("CRM Sistem")
        self.setGeometry(50, 50, 1600, 1000)
        self.setStyleSheet(STYLESHEET)
        
        self.giris_ekrani_ac()
    
    def giris_ekrani_ac(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("CRM Giris")
        dialog.setModal(True)
        dialog.setGeometry(300, 300, 400, 300)
        dialog.setStyleSheet(STYLESHEET)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        title = QLabel("CRM")
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Kullanici Adi"))
        kullanici_input = QLineEdit()
        layout.addWidget(kullanici_input)
        
        layout.addWidget(QLabel("Sifre"))
        sifre_input = QLineEdit()
        sifre_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(sifre_input)
        
        def giris_yap():
            kadi = kullanici_input.text()
            sifre = sifre_input.text()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT * FROM kullanicilar WHERE kullanici_adi=? AND sifre=?', (kadi, sifre))
                kullanici = cursor.fetchone()
            
            if kullanici:
                self.kullanici = dict(kullanici)
                dialog.accept()
                self.ui_baslat()
            else:
                QMessageBox.warning(dialog, "Hata", "Kullanici adi veya sifre yanlis!")
        
        giris_btn = QPushButton("Giris Yap")
        giris_btn.clicked.connect(giris_yap)
        layout.addWidget(giris_btn)
        
        layout.addWidget(QLabel("Demo: admin / admin123"))
        
        dialog.setLayout(layout)
        if dialog.exec_() != QDialog.Accepted:
            sys.exit()
    
    def ui_baslat(self):
        central = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        header = QFrame()
        header.setStyleSheet(f"background-color: {COLORS['bg_secondary']}; border-bottom: 1px solid {COLORS['border']};")
        header.setFixedHeight(60)
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(30, 0, 30, 0)
        
        title = QLabel(f"CRM | Hos geldiniz {self.kullanici['ad']} {self.kullanici['soyad']}")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setStyleSheet(f"color: {COLORS['primary']};")
        
        cikis_btn = QPushButton("Cikis")
        cikis_btn.setFixedWidth(100)
        cikis_btn.clicked.connect(self.cikis_yap)
        
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(cikis_btn)
        header.setLayout(header_layout)
        
        self.tabs = QTabWidget()
        self.tab_dashboard = self.dashboard_tab_olustur()
        self.tab_musteriler = self.musteri_tab_olustur()
        self.tab_satislar = self.satis_tab_olustur()
        self.tab_destekler = self.destek_tab_olustur()
        self.tab_oduller = self.oduller_tab_olustur()
        self.tab_analitik = self.analitik_tab_olustur()
        self.tab_email = self.email_tab_olustur()
        self.tab_kampanya = self.kampanya_tab_olustur()
        self.tab_firsat = self.firsat_tab_olustur()
        self.tab_takvim = self.takvim_tab_olustur()
        self.tab_profil = self.profil_tab_olustur()
        self.tab_raporlar = self.raporlar_tab_olustur()
        
        self.tabs.addTab(self.tab_dashboard, "Dashboard")
        self.tabs.addTab(self.tab_musteriler, "Musteriler")
        self.tabs.addTab(self.tab_satislar, "Satislar")
        self.tabs.addTab(self.tab_destekler, "Destek")
        self.tabs.addTab(self.tab_oduller, "Oduller")
        self.tabs.addTab(self.tab_analitik, "Analitik")
        self.tabs.addTab(self.tab_email, "Email")
        self.tabs.addTab(self.tab_kampanya, "Kampanyalar")
        self.tabs.addTab(self.tab_firsat, "Firsatlar")
        self.tabs.addTab(self.tab_takvim, "Takvim")
        self.tabs.addTab(self.tab_profil, "Profil")
        self.tabs.addTab(self.tab_raporlar, "Raporlar")
        
        main_layout.addWidget(header)
        main_layout.addWidget(self.tabs)
        
        central.setLayout(main_layout)
        self.setCentralWidget(central)
        
        self.veriyi_yenile()
    
    def dashboard_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.kpi_gelir = self.kpi_kart("Toplam Gelir", "0 TRY", COLORS['success'])
        self.kpi_musteri = self.kpi_kart("Aktif Musteri", "0", COLORS['primary'])
        self.kpi_satis = self.kpi_kart("Satis Sayisi", "0", COLORS['warning'])
        self.kpi_destek = self.kpi_kart("Acik Destek", "0", COLORS['danger'])
        
        kpi_layout.addWidget(self.kpi_gelir)
        kpi_layout.addWidget(self.kpi_musteri)
        kpi_layout.addWidget(self.kpi_satis)
        kpi_layout.addWidget(self.kpi_destek)
        
        layout.addLayout(kpi_layout)
        
        grafik_layout = QHBoxLayout()
        grafik_layout.setSpacing(15)
        
        self.gelir_grafik = self.grafik_widget()
        self.durum_grafik = self.grafik_widget()
        
        gelir_frame = QFrame()
        gelir_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        gelir_vlayout = QVBoxLayout()
        gelir_baslik = QLabel("Aylik Gelir Trendi")
        gelir_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        gelir_vlayout.addWidget(gelir_baslik)
        gelir_vlayout.addWidget(self.gelir_grafik)
        gelir_frame.setLayout(gelir_vlayout)
        
        durum_frame = QFrame()
        durum_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        durum_vlayout = QVBoxLayout()
        durum_baslik = QLabel("Musteri Durumu")
        durum_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        durum_vlayout.addWidget(durum_baslik)
        durum_vlayout.addWidget(self.durum_grafik)
        durum_frame.setLayout(durum_vlayout)
        
        grafik_layout.addWidget(gelir_frame)
        grafik_layout.addWidget(durum_frame)
        
        layout.addLayout(grafik_layout)
        
        layout.addWidget(QLabel("Son 10 Satis"))
        
        self.dashboard_tablo = QTableWidget()
        self.dashboard_tablo.setColumnCount(6)
        self.dashboard_tablo.setHorizontalHeaderLabels(["ID", "Musteri", "Urun", "Tutar", "Durum", "Tarih"])
        self.dashboard_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.dashboard_tablo.setMaximumHeight(250)
        
        layout.addWidget(self.dashboard_tablo)
        
        widget.setLayout(layout)
        
        self.dashboard_timer = QTimer()
        self.dashboard_timer.timeout.connect(self.dashboard_yenile)
        self.dashboard_timer.start(5000)
        
        return widget
    
    def kpi_kart(self, baslik, deger, renk):
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background-color: {COLORS['bg_secondary']};
                border: 1px solid {COLORS['border']};
                border-radius: 12px;
                padding: 20px;
            }}
        """)
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        baslik_label = QLabel(baslik)
        baslik_label.setFont(QFont("Arial", 11))
        baslik_label.setStyleSheet(f"color: {COLORS['text_sec']};")
        
        deger_label = QLabel(deger)
        deger_label.setFont(QFont("Arial", 24, QFont.Bold))
        deger_label.setStyleSheet(f"color: {renk};")
        
        layout.addWidget(baslik_label)
        layout.addWidget(deger_label)
        layout.addStretch()
        
        frame.setLayout(layout)
        frame.deger_label = deger_label
        
        return frame
    
    def grafik_widget(self):
        fig = Figure(figsize=(5, 3), dpi=100, facecolor=COLORS['bg_secondary'])
        canvas = FigureCanvas(fig)
        canvas.setStyleSheet(f"background-color: {COLORS['bg_secondary']};")
        canvas.fig = fig
        return canvas
    
    def dashboard_yenile(self):
        musteriler = self.db.musterileri_getir()
        satislar = self.db.satislari_getir()
        destekler = self.db.destekleri_getir()
        
        toplam_gelir = sum(s['toplam'] for s in satislar)
        aktif_musteri = sum(1 for m in musteriler if m['durum'] == 'Aktif')
        satis_sayisi = len(satislar)
        acik_destek = sum(1 for d in destekler if d['durum'] == 'Acik')
        
        self.kpi_gelir.deger_label.setText(f"{toplam_gelir:,.0f} TRY")
        self.kpi_musteri.deger_label.setText(str(aktif_musteri))
        self.kpi_satis.deger_label.setText(str(satis_sayisi))
        self.kpi_destek.deger_label.setText(str(acik_destek))
        
        self.gelir_grafik.fig.clear()
        ax1 = self.gelir_grafik.fig.add_subplot(111)
        
        aylar = {}
        for satis in satislar:
            ay = satis['tarih'][:7]
            aylar[ay] = aylar.get(ay, 0) + satis['toplam']
        
        if aylar:
            months = sorted(aylar.keys())[-6:]
            values = [aylar[m] for m in months]
            ax1.bar(range(len(months)), values, color=COLORS['primary'], alpha=0.8)
            ax1.set_xticks(range(len(months)))
            ax1.set_xticklabels([m[5:] for m in months], color=COLORS['text_sec'])
            ax1.set_ylabel('Gelir (TRY)', color=COLORS['text_sec'])
            ax1.tick_params(colors=COLORS['text_sec'])
        
        ax1.set_facecolor(COLORS['bg_secondary'])
        ax1.spines['bottom'].set_color(COLORS['border'])
        ax1.spines['left'].set_color(COLORS['border'])
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        ax1.grid(True, alpha=0.1)
        
        self.gelir_grafik.draw()
        
        self.durum_grafik.fig.clear()
        ax2 = self.durum_grafik.fig.add_subplot(111)
        
        durum_say = {}
        for musteri in musteriler:
            durum = musteri['durum']
            durum_say[durum] = durum_say.get(durum, 0) + 1
        
        if durum_say:
            labels = list(durum_say.keys())
            sizes = list(durum_say.values())
            colors_pie = [COLORS['success'] if l == 'Aktif' else COLORS['danger'] if l == 'Pasif' else COLORS['warning'] for l in labels]
            ax2.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors_pie, startangle=90)
        
        for text in ax2.texts:
            text.set_color(COLORS['text_main'])
        
        self.durum_grafik.draw()
        
        self.dashboard_tablo.setRowCount(0)
        for satis in sorted(satislar, key=lambda x: x['id'], reverse=True)[:10]:
            row = self.dashboard_tablo.rowCount()
            self.dashboard_tablo.insertRow(row)
            
            self.dashboard_tablo.setItem(row, 0, QTableWidgetItem(str(satis['id'])))
            self.dashboard_tablo.setItem(row, 1, QTableWidgetItem(satis['musteri_adi']))
            self.dashboard_tablo.setItem(row, 2, QTableWidgetItem(satis['urun']))
            self.dashboard_tablo.setItem(row, 3, QTableWidgetItem(f"{satis['tutar']:,.0f} TRY"))
            
            durum_item = QTableWidgetItem(satis['durum'])
            if satis['durum'] == 'Tamamlandi':
                durum_item.setForeground(QColor(COLORS['success']))
            elif satis['durum'] == 'Isleniysor':
                durum_item.setForeground(QColor(COLORS['warning']))
            else:
                durum_item.setForeground(QColor(COLORS['danger']))
            self.dashboard_tablo.setItem(row, 4, durum_item)
            
            self.dashboard_tablo.setItem(row, 5, QTableWidgetItem(satis['tarih']))
    
    def musteri_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        toolbar = QHBoxLayout()
        
        self.musteri_arama = QLineEdit()
        self.musteri_arama.setPlaceholderText("Musteri ara...")
        self.musteri_arama.setFixedWidth(300)
        self.musteri_arama.textChanged.connect(self.musterileri_ara)
        
        ekle_btn = QPushButton("Musteri Ekle")
        ekle_btn.clicked.connect(self.musteri_ekle)
        
        duzenle_btn = QPushButton("Duzenle")
        duzenle_btn.clicked.connect(self.musteri_duzenle)
        
        sil_btn = QPushButton("Sil")
        sil_btn.clicked.connect(self.musteri_sil)
        
        toolbar.addWidget(self.musteri_arama)
        toolbar.addStretch()
        toolbar.addWidget(ekle_btn)
        toolbar.addWidget(duzenle_btn)
        toolbar.addWidget(sil_btn)
        
        self.musteri_tablo = QTableWidget()
        self.musteri_tablo.setColumnCount(7)
        self.musteri_tablo.setHorizontalHeaderLabels(["ID", "Ad", "Email", "Telefon", "Sirket", "Durum", "Toplam"])
        self.musteri_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addLayout(toolbar)
        layout.addWidget(self.musteri_tablo)
        
        widget.setLayout(layout)
        return widget
    
    def satis_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        toolbar = QHBoxLayout()
        
        self.satis_arama = QLineEdit()
        self.satis_arama.setPlaceholderText("Satis ara...")
        self.satis_arama.setFixedWidth(300)
        self.satis_arama.textChanged.connect(self.satislari_ara)
        
        ekle_btn = QPushButton("Satis Ekle")
        ekle_btn.clicked.connect(self.satis_ekle)
        
        duzenle_btn = QPushButton("Duzenle")
        duzenle_btn.clicked.connect(self.satis_duzenle)
        
        sil_btn = QPushButton("Sil")
        sil_btn.clicked.connect(self.satis_sil)
        
        toolbar.addWidget(self.satis_arama)
        toolbar.addStretch()
        toolbar.addWidget(ekle_btn)
        toolbar.addWidget(duzenle_btn)
        toolbar.addWidget(sil_btn)
        
        self.satis_tablo = QTableWidget()
        self.satis_tablo.setColumnCount(7)
        self.satis_tablo.setHorizontalHeaderLabels(["ID", "Musteri", "Urun", "Tutar", "Adet", "Toplam", "Durum"])
        self.satis_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addLayout(toolbar)
        layout.addWidget(self.satis_tablo)
        
        widget.setLayout(layout)
        return widget
    
    def destek_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        toolbar = QHBoxLayout()
        
        self.destek_arama = QLineEdit()
        self.destek_arama.setPlaceholderText("Destek ara...")
        self.destek_arama.setFixedWidth(300)
        self.destek_arama.textChanged.connect(self.destekleri_ara)
        
        ekle_btn = QPushButton("Destek Ekle")
        ekle_btn.clicked.connect(self.destek_ekle)
        
        duzenle_btn = QPushButton("Duzenle")
        duzenle_btn.clicked.connect(self.destek_duzenle)
        
        sil_btn = QPushButton("Sil")
        sil_btn.clicked.connect(self.destek_sil)
        
        toolbar.addWidget(self.destek_arama)
        toolbar.addStretch()
        toolbar.addWidget(ekle_btn)
        toolbar.addWidget(duzenle_btn)
        toolbar.addWidget(sil_btn)
        
        self.destek_tablo = QTableWidget()
        self.destek_tablo.setColumnCount(7)
        self.destek_tablo.setHorizontalHeaderLabels(["ID", "Musteri", "Konu", "Aciklama", "Oncelik", "Durum", "Tarih"])
        self.destek_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addLayout(toolbar)
        layout.addWidget(self.destek_tablo)
        
        widget.setLayout(layout)
        return widget
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 10: MUSTERI PROFIL SAYFASI (360 DERECE GORUNUM)
    # ─────────────────────────────────────────────────────────────────────────
    
    def profil_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # ÜST: Müşteri Seçici
        secici_layout = QHBoxLayout()
        secici_layout.addWidget(QLabel("Musteri Sec:"))
        
        self.profil_musteri_combo = QComboBox()
        self.profil_musteri_combo.setMinimumWidth(400)
        self.profil_musteri_combo.currentIndexChanged.connect(self.profil_yukle)
        secici_layout.addWidget(self.profil_musteri_combo)
        
        yenile_btn = QPushButton("Yenile")
        yenile_btn.clicked.connect(self.profil_yukle)
        secici_layout.addWidget(yenile_btn)
        
        secici_layout.addStretch()
        layout.addLayout(secici_layout)
        
        # MUSTERI BILGI KARTI (Üst Header)
        self.profil_header = QFrame()
        self.profil_header.setStyleSheet(f"""
            QFrame {{
                background-color: {COLORS['bg_secondary']};
                border: 1px solid {COLORS['border']};
                border-radius: 12px;
                padding: 20px;
            }}
        """)
        self.profil_header.setMinimumHeight(150)
        
        header_layout = QHBoxLayout()
        
        # Sol: İsim, Şirket, Durum
        sol_bilgi = QVBoxLayout()
        self.profil_ad_label = QLabel("Musteri Sec")
        self.profil_ad_label.setFont(QFont("Arial", 22, QFont.Bold))
        
        self.profil_sirket_label = QLabel("")
        self.profil_sirket_label.setFont(QFont("Arial", 12))
        self.profil_sirket_label.setStyleSheet(f"color: {COLORS['text_sec']};")
        
        self.profil_durum_label = QLabel("")
        self.profil_durum_label.setFont(QFont("Arial", 11, QFont.Bold))
        
        sol_bilgi.addWidget(self.profil_ad_label)
        sol_bilgi.addWidget(self.profil_sirket_label)
        sol_bilgi.addWidget(self.profil_durum_label)
        sol_bilgi.addStretch()
        
        # Orta: İletişim
        orta_bilgi = QVBoxLayout()
        self.profil_email_label = QLabel("")
        self.profil_email_label.setStyleSheet(f"color: {COLORS['text_sec']};")
        
        self.profil_tel_label = QLabel("")
        self.profil_tel_label.setStyleSheet(f"color: {COLORS['text_sec']};")
        
        self.profil_katilim_label = QLabel("")
        self.profil_katilim_label.setStyleSheet(f"color: {COLORS['text_sec']};")
        
        orta_bilgi.addWidget(self.profil_email_label)
        orta_bilgi.addWidget(self.profil_tel_label)
        orta_bilgi.addWidget(self.profil_katilim_label)
        orta_bilgi.addStretch()
        
        # Sag: Aksiyon Butonlari
        sag_aksiyon = QVBoxLayout()
        
        duzenle_btn = QPushButton("Musteri Duzenle")
        duzenle_btn.clicked.connect(self.profil_musteri_duzenle)
        
        satis_ekle_btn = QPushButton("Satis Ekle")
        satis_ekle_btn.clicked.connect(self.profil_satis_ekle)
        
        etkinlik_ekle_btn = QPushButton("Etkinlik Olustur")
        etkinlik_ekle_btn.clicked.connect(self.profil_etkinlik_ekle)
        
        sag_aksiyon.addWidget(duzenle_btn)
        sag_aksiyon.addWidget(satis_ekle_btn)
        sag_aksiyon.addWidget(etkinlik_ekle_btn)
        sag_aksiyon.addStretch()
        
        header_layout.addLayout(sol_bilgi, 2)
        header_layout.addLayout(orta_bilgi, 2)
        header_layout.addLayout(sag_aksiyon, 1)
        
        self.profil_header.setLayout(header_layout)
        layout.addWidget(self.profil_header)
        
        # KPI Kartlari (Profil ozet)
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.profil_kpi_satis = self.kpi_kart("Toplam Satis", "0", COLORS['success'])
        self.profil_kpi_harcama = self.kpi_kart("Toplam Harcama", "0 TRY", COLORS['primary'])
        self.profil_kpi_puan = self.kpi_kart("Puan", "0", COLORS['warning'])
        self.profil_kpi_destek = self.kpi_kart("Destek Talebi", "0", COLORS['danger'])
        self.profil_kpi_firsat = self.kpi_kart("Aktif Firsat", "0", COLORS['secondary'])
        
        kpi_layout.addWidget(self.profil_kpi_satis)
        kpi_layout.addWidget(self.profil_kpi_harcama)
        kpi_layout.addWidget(self.profil_kpi_puan)
        kpi_layout.addWidget(self.profil_kpi_destek)
        kpi_layout.addWidget(self.profil_kpi_firsat)
        
        layout.addLayout(kpi_layout)
        
        # ORTA: 2 Panel (Satin Alma Grafiği + Aktivite Zaman Çizelgesi)
        orta_layout = QHBoxLayout()
        orta_layout.setSpacing(15)
        
        # SOL: Satin Alma Grafigi
        self.profil_satis_grafik = self.grafik_widget()
        sol_frame = QFrame()
        sol_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        sol_vlayout = QVBoxLayout()
        sol_baslik = QLabel("Satin Alma Trendi")
        sol_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        sol_vlayout.addWidget(sol_baslik)
        sol_vlayout.addWidget(self.profil_satis_grafik)
        sol_frame.setLayout(sol_vlayout)
        
        # SAG: Aktivite Skorlari (Pie Chart)
        self.profil_aktivite_grafik = self.grafik_widget()
        sag_frame = QFrame()
        sag_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        sag_vlayout = QVBoxLayout()
        sag_baslik = QLabel("Aktivite Dagilimi")
        sag_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        sag_vlayout.addWidget(sag_baslik)
        sag_vlayout.addWidget(self.profil_aktivite_grafik)
        sag_frame.setLayout(sag_vlayout)
        
        orta_layout.addWidget(sol_frame)
        orta_layout.addWidget(sag_frame)
        
        layout.addLayout(orta_layout)
        
        # ALT: 3 Tablo (Satışlar, Destek, Etkinlikler)
        alt_layout = QHBoxLayout()
        alt_layout.setSpacing(15)
        
        # Sol: Satıs Geçmişi
        sol_alt = QWidget()
        sol_alt_layout = QVBoxLayout()
        sol_alt_layout.setContentsMargins(0, 0, 0, 0)
        sol_alt_baslik = QLabel("Satin Alma Gecmisi")
        sol_alt_baslik.setFont(QFont("Arial", 11, QFont.Bold))
        sol_alt_layout.addWidget(sol_alt_baslik)
        
        self.profil_satis_tablo = QTableWidget()
        self.profil_satis_tablo.setColumnCount(4)
        self.profil_satis_tablo.setHorizontalHeaderLabels(["Tarih", "Urun", "Tutar", "Durum"])
        self.profil_satis_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.profil_satis_tablo.setMaximumHeight(220)
        sol_alt_layout.addWidget(self.profil_satis_tablo)
        sol_alt.setLayout(sol_alt_layout)
        
        # Orta: Destek Talepleri
        orta_alt = QWidget()
        orta_alt_layout = QVBoxLayout()
        orta_alt_layout.setContentsMargins(0, 0, 0, 0)
        orta_alt_baslik = QLabel("Destek Talepleri")
        orta_alt_baslik.setFont(QFont("Arial", 11, QFont.Bold))
        orta_alt_layout.addWidget(orta_alt_baslik)
        
        self.profil_destek_tablo = QTableWidget()
        self.profil_destek_tablo.setColumnCount(4)
        self.profil_destek_tablo.setHorizontalHeaderLabels(["Tarih", "Konu", "Oncelik", "Durum"])
        self.profil_destek_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.profil_destek_tablo.setMaximumHeight(220)
        orta_alt_layout.addWidget(self.profil_destek_tablo)
        orta_alt.setLayout(orta_alt_layout)
        
        # Sağ: Etkinlikler
        sag_alt = QWidget()
        sag_alt_layout = QVBoxLayout()
        sag_alt_layout.setContentsMargins(0, 0, 0, 0)
        sag_alt_baslik = QLabel("Etkinlikler")
        sag_alt_baslik.setFont(QFont("Arial", 11, QFont.Bold))
        sag_alt_layout.addWidget(sag_alt_baslik)
        
        self.profil_etkinlik_tablo = QTableWidget()
        self.profil_etkinlik_tablo.setColumnCount(4)
        self.profil_etkinlik_tablo.setHorizontalHeaderLabels(["Tarih", "Baslik", "Tip", "Durum"])
        self.profil_etkinlik_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.profil_etkinlik_tablo.setMaximumHeight(220)
        sag_alt_layout.addWidget(self.profil_etkinlik_tablo)
        sag_alt.setLayout(sag_alt_layout)
        
        alt_layout.addWidget(sol_alt)
        alt_layout.addWidget(orta_alt)
        alt_layout.addWidget(sag_alt)
        
        layout.addLayout(alt_layout)
        
        widget.setLayout(layout)
        return widget
    
    def profil_musteri_duzenle(self):
        musteri_id = self.profil_musteri_combo.currentData()
        if not musteri_id:
            return
        
        musteriler = self.db.musterileri_getir()
        musteri = next((m for m in musteriler if m['id'] == musteri_id), None)
        
        dialog = MusteriDialog(self, musteri)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.musteri_guncelle(musteri_id, veri['ad'], veri['email'], veri['telefon'], veri['sirket'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Musteri guncellendi!")
            self.veriyi_yenile()
    
    def profil_satis_ekle(self):
        musteri_id = self.profil_musteri_combo.currentData()
        if not musteri_id:
            QMessageBox.warning(self, "Uyari", "Once musteri secin!")
            return
        
        dialog = SatisDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.satis_ekle(musteri_id, veri['urun'], veri['tutar'], veri['adet'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Satis eklendi!")
            self.veriyi_yenile()
    
    def profil_etkinlik_ekle(self):
        musteri_id = self.profil_musteri_combo.currentData()
        if not musteri_id:
            QMessageBox.warning(self, "Uyari", "Once musteri secin!")
            return
        
        musteriler = self.db.musterileri_getir()
        dialog = EtkinlikDialog(self, musteriler)
        # Müşteri seçimini otomatik yap
        for i in range(dialog.musteri_combo.count()):
            if dialog.musteri_combo.itemData(i) == musteri_id:
                dialog.musteri_combo.setCurrentIndex(i)
                break
        
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.etkinlik_ekle(
                veri['musteri_id'], veri['baslik'], veri['aciklama'], veri['tip'],
                veri['oncelik'], veri['tarih'], veri['saat'], veri['sure'],
                veri['konum'], veri['durum'], veri['hatirlatici']
            )
            QMessageBox.information(self, "Basarili", "Etkinlik eklendi!")
            self.veriyi_yenile()
    
    def profil_combo_doldur(self):
        if not hasattr(self, 'profil_musteri_combo'):
            return
        
        musteriler = self.db.musterileri_getir()
        mevcut_id = self.profil_musteri_combo.currentData()
        
        self.profil_musteri_combo.blockSignals(True)
        self.profil_musteri_combo.clear()
        self.profil_musteri_combo.addItem("--- Musteri Sec ---", None)
        
        for m in musteriler:
            self.profil_musteri_combo.addItem(f"{m['ad']} - {m['sirket'] or 'Bireysel'} ({m['email']})", m['id'])
        
        # Önceki seçimi geri yükle
        if mevcut_id:
            for i in range(self.profil_musteri_combo.count()):
                if self.profil_musteri_combo.itemData(i) == mevcut_id:
                    self.profil_musteri_combo.setCurrentIndex(i)
                    break
        
        self.profil_musteri_combo.blockSignals(False)
    
    def profil_yukle(self):
        if not hasattr(self, 'profil_musteri_combo'):
            return
        
        musteri_id = self.profil_musteri_combo.currentData()
        if not musteri_id:
            self.profil_ad_label.setText("Musteri Sec")
            self.profil_sirket_label.setText("")
            self.profil_durum_label.setText("")
            self.profil_email_label.setText("")
            self.profil_tel_label.setText("")
            self.profil_katilim_label.setText("")
            
            self.profil_kpi_satis.deger_label.setText("0")
            self.profil_kpi_harcama.deger_label.setText("0 TRY")
            self.profil_kpi_puan.deger_label.setText("0")
            self.profil_kpi_destek.deger_label.setText("0")
            self.profil_kpi_firsat.deger_label.setText("0")
            
            self.profil_satis_tablo.setRowCount(0)
            self.profil_destek_tablo.setRowCount(0)
            self.profil_etkinlik_tablo.setRowCount(0)
            
            self.profil_satis_grafik.fig.clear()
            self.profil_satis_grafik.draw()
            self.profil_aktivite_grafik.fig.clear()
            self.profil_aktivite_grafik.draw()
            return
        
        # Müşteri bilgilerini al
        musteriler = self.db.musterileri_getir()
        musteri = next((m for m in musteriler if m['id'] == musteri_id), None)
        if not musteri:
            return
        
        # ─── HEADER BİLGİLERİ ───
        self.profil_ad_label.setText(musteri['ad'])
        self.profil_sirket_label.setText(musteri['sirket'] or "Bireysel Musteri")
        
        durum = musteri['durum']
        if durum == 'Aktif':
            self.profil_durum_label.setText(f"● AKTIF")
            self.profil_durum_label.setStyleSheet(f"color: {COLORS['success']};")
        elif durum == 'Pasif':
            self.profil_durum_label.setText(f"● PASIF")
            self.profil_durum_label.setStyleSheet(f"color: {COLORS['danger']};")
        else:
            self.profil_durum_label.setText(f"● BEKLEMEDE")
            self.profil_durum_label.setStyleSheet(f"color: {COLORS['warning']};")
        
        self.profil_email_label.setText(f"📧 {musteri['email']}")
        self.profil_tel_label.setText(f"📞 {musteri['telefon'] or 'Belirtilmemis'}")
        self.profil_katilim_label.setText(f"📅 Katilim: {musteri['katilim_tarihi']}")
        
        # ─── İLGİLİ VERİLERİ ÇEK ───
        satislar = self.db.satislari_getir()
        musteri_satislar = [s for s in satislar if s['musteri_id'] == musteri_id]
        
        destekler = self.db.destekleri_getir()
        musteri_destekler = [d for d in destekler if d['musteri_id'] == musteri_id]
        
        etkinlikler = self.db.etkinlikleri_getir()
        musteri_etkinlikler = [e for e in etkinlikler if e['musteri_id'] == musteri_id]
        
        firsatlar = self.db.firsatlari_getir()
        musteri_firsatlar = [f for f in firsatlar if f['musteri_id'] == musteri_id]
        aktif_firsatlar = [f for f in musteri_firsatlar if f['asama'] not in ['Kazanildi', 'Kaybedildi']]
        
        puan_data = self.db.musteri_puan_getir(musteri_id)
        puan = puan_data['puan'] if puan_data else 0
        
        # ─── KPI'LARI GÜNCELLE ───
        toplam_harcama = sum(s['toplam'] for s in musteri_satislar if s['durum'] == 'Tamamlandi')
        
        self.profil_kpi_satis.deger_label.setText(str(len(musteri_satislar)))
        self.profil_kpi_harcama.deger_label.setText(f"{toplam_harcama:,.0f} TRY")
        self.profil_kpi_puan.deger_label.setText(f"{puan:,}")
        self.profil_kpi_destek.deger_label.setText(str(len(musteri_destekler)))
        self.profil_kpi_firsat.deger_label.setText(str(len(aktif_firsatlar)))
        
        # ─── SATIN ALMA TRENDİ GRAFIGI ───
        self.profil_satis_grafik.fig.clear()
        ax1 = self.profil_satis_grafik.fig.add_subplot(111)
        
        if musteri_satislar:
            aylar = {}
            for s in musteri_satislar:
                ay = s['tarih'][:7]
                aylar[ay] = aylar.get(ay, 0) + s['toplam']
            
            sorted_aylar = sorted(aylar.keys())[-12:]
            degerler = [aylar[a] for a in sorted_aylar]
            
            ax1.plot(range(len(sorted_aylar)), degerler, color=COLORS['primary'], linewidth=2, marker='o', markersize=8)
            ax1.fill_between(range(len(sorted_aylar)), degerler, alpha=0.3, color=COLORS['primary'])
            ax1.set_xticks(range(len(sorted_aylar)))
            ax1.set_xticklabels([a[5:] for a in sorted_aylar], color=COLORS['text_sec'], rotation=45)
            ax1.set_ylabel('Tutar (TRY)', color=COLORS['text_sec'])
            ax1.tick_params(colors=COLORS['text_sec'])
        else:
            ax1.text(0.5, 0.5, 'Veri Yok', ha='center', va='center', 
                     color=COLORS['text_sec'], fontsize=14, transform=ax1.transAxes)
        
        ax1.set_facecolor(COLORS['bg_secondary'])
        ax1.spines['bottom'].set_color(COLORS['border'])
        ax1.spines['left'].set_color(COLORS['border'])
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        ax1.grid(True, alpha=0.1)
        
        self.profil_satis_grafik.fig.tight_layout()
        self.profil_satis_grafik.draw()
        
        # ─── AKTIVITE DAĞILIMI GRAFIGI ───
        self.profil_aktivite_grafik.fig.clear()
        ax2 = self.profil_aktivite_grafik.fig.add_subplot(111)
        
        aktivite = {
            'Satis': len(musteri_satislar),
            'Destek': len(musteri_destekler),
            'Etkinlik': len(musteri_etkinlikler),
            'Firsat': len(musteri_firsatlar)
        }
        
        # Sadece 0'dan büyük olanları göster
        aktivite = {k: v for k, v in aktivite.items() if v > 0}
        
        if aktivite:
            labels = list(aktivite.keys())
            sizes = list(aktivite.values())
            colors_pie = [COLORS['success'], COLORS['danger'], COLORS['warning'], COLORS['primary']]
            ax2.pie(sizes, labels=labels, autopct='%1.0f%%', colors=colors_pie[:len(labels)], startangle=90)
        else:
            ax2.text(0.5, 0.5, 'Aktivite Yok', ha='center', va='center', 
                     color=COLORS['text_sec'], fontsize=14, transform=ax2.transAxes)
        
        for text in ax2.texts:
            text.set_color(COLORS['text_main'])
        
        self.profil_aktivite_grafik.draw()
        
        # ─── SATIS TABLOSU ───
        self.profil_satis_tablo.setRowCount(0)
        for s in musteri_satislar[:20]:
            row = self.profil_satis_tablo.rowCount()
            self.profil_satis_tablo.insertRow(row)
            
            self.profil_satis_tablo.setItem(row, 0, QTableWidgetItem(s['tarih']))
            self.profil_satis_tablo.setItem(row, 1, QTableWidgetItem(s['urun']))
            self.profil_satis_tablo.setItem(row, 2, QTableWidgetItem(f"{s['toplam']:,.0f} TRY"))
            
            durum_item = QTableWidgetItem(s['durum'])
            if s['durum'] == 'Tamamlandi':
                durum_item.setForeground(QColor(COLORS['success']))
            elif s['durum'] == 'Isleniysor':
                durum_item.setForeground(QColor(COLORS['warning']))
            else:
                durum_item.setForeground(QColor(COLORS['danger']))
            self.profil_satis_tablo.setItem(row, 3, durum_item)
        
        # ─── DESTEK TABLOSU ───
        self.profil_destek_tablo.setRowCount(0)
        for d in musteri_destekler[:20]:
            row = self.profil_destek_tablo.rowCount()
            self.profil_destek_tablo.insertRow(row)
            
            self.profil_destek_tablo.setItem(row, 0, QTableWidgetItem(d['tarih']))
            self.profil_destek_tablo.setItem(row, 1, QTableWidgetItem(d['konu']))
            
            oncelik_item = QTableWidgetItem(d['oncelik'])
            if d['oncelik'] == 'Acil':
                oncelik_item.setForeground(QColor(COLORS['danger']))
            elif d['oncelik'] == 'Yuksek':
                oncelik_item.setForeground(QColor(COLORS['warning']))
            self.profil_destek_tablo.setItem(row, 2, oncelik_item)
            
            durum_item = QTableWidgetItem(d['durum'])
            if d['durum'] == 'Cozuldu':
                durum_item.setForeground(QColor(COLORS['success']))
            elif d['durum'] == 'Acik':
                durum_item.setForeground(QColor(COLORS['danger']))
            self.profil_destek_tablo.setItem(row, 3, durum_item)
        
        # ─── ETKINLIK TABLOSU ───
        self.profil_etkinlik_tablo.setRowCount(0)
        for e in musteri_etkinlikler[:20]:
            row = self.profil_etkinlik_tablo.rowCount()
            self.profil_etkinlik_tablo.insertRow(row)
            
            self.profil_etkinlik_tablo.setItem(row, 0, QTableWidgetItem(e['tarih']))
            self.profil_etkinlik_tablo.setItem(row, 1, QTableWidgetItem(e['baslik']))
            self.profil_etkinlik_tablo.setItem(row, 2, QTableWidgetItem(e['tip']))
            
            durum_item = QTableWidgetItem(e['durum'])
            if e['durum'] == 'Tamamlandi':
                durum_item.setForeground(QColor(COLORS['success']))
            elif e['durum'] == 'Iptal':
                durum_item.setForeground(QColor(COLORS['danger']))
            elif e['durum'] == 'Ertelendi':
                durum_item.setForeground(QColor(COLORS['warning']))
            self.profil_etkinlik_tablo.setItem(row, 3, durum_item)
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 9: TAKVIM & PLANLAMA
    # ─────────────────────────────────────────────────────────────────────────
    
    def takvim_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # KPI Kartlari
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.kpi_bugun_etkinlik = self.kpi_kart("Bugunki Etkinlik", "0", COLORS['primary'])
        self.kpi_bu_hafta = self.kpi_kart("Bu Hafta", "0", COLORS['warning'])
        self.kpi_yaklasan = self.kpi_kart("Acil/Yaklaşan", "0", COLORS['danger'])
        self.kpi_tamamlanan = self.kpi_kart("Tamamlanan", "0", COLORS['success'])
        
        kpi_layout.addWidget(self.kpi_bugun_etkinlik)
        kpi_layout.addWidget(self.kpi_bu_hafta)
        kpi_layout.addWidget(self.kpi_yaklasan)
        kpi_layout.addWidget(self.kpi_tamamlanan)
        
        layout.addLayout(kpi_layout)
        
        # Aksiyon Butonlari
        aksiyon_layout = QHBoxLayout()
        
        ekle_btn = QPushButton("Yeni Etkinlik")
        ekle_btn.clicked.connect(self.etkinlik_ekle)
        
        duzenle_btn = QPushButton("Duzenle")
        duzenle_btn.clicked.connect(self.etkinlik_duzenle)
        
        sil_btn = QPushButton("Sil")
        sil_btn.clicked.connect(self.etkinlik_sil)
        
        tamamla_btn = QPushButton("Tamamla")
        tamamla_btn.clicked.connect(lambda: self.etkinlik_durum_degistir("Tamamlandi"))
        
        iptal_btn = QPushButton("Iptal Et")
        iptal_btn.clicked.connect(lambda: self.etkinlik_durum_degistir("Iptal"))
        
        ertele_btn = QPushButton("Ertele")
        ertele_btn.clicked.connect(lambda: self.etkinlik_durum_degistir("Ertelendi"))
        
        aksiyon_layout.addWidget(ekle_btn)
        aksiyon_layout.addWidget(duzenle_btn)
        aksiyon_layout.addWidget(sil_btn)
        aksiyon_layout.addStretch()
        aksiyon_layout.addWidget(tamamla_btn)
        aksiyon_layout.addWidget(ertele_btn)
        aksiyon_layout.addWidget(iptal_btn)
        
        layout.addLayout(aksiyon_layout)
        
        # Ana panel: Takvim + Etkinlik Listesi
        ana_layout = QHBoxLayout()
        ana_layout.setSpacing(15)
        
        # SOL: TAKVIM
        sol_widget = QWidget()
        sol_layout = QVBoxLayout()
        sol_layout.setContentsMargins(0, 0, 0, 0)
        
        sol_baslik = QLabel("Takvim")
        sol_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        sol_layout.addWidget(sol_baslik)
        
        self.takvim_widget = QCalendarWidget()
        self.takvim_widget.setStyleSheet(f"""
            QCalendarWidget QAbstractItemView:enabled {{
                background-color: {COLORS['bg_tertiary']};
                color: {COLORS['text_main']};
                selection-background-color: {COLORS['primary']};
                selection-color: white;
            }}
            QCalendarWidget QToolButton {{
                color: {COLORS['text_main']};
                background-color: {COLORS['bg_secondary']};
            }}
            QCalendarWidget QWidget {{
                background-color: {COLORS['bg_main']};
            }}
            QCalendarWidget QMenu {{
                background-color: {COLORS['bg_secondary']};
                color: {COLORS['text_main']};
            }}
            QCalendarWidget QSpinBox {{
                background-color: {COLORS['bg_secondary']};
                color: {COLORS['text_main']};
            }}
        """)
        self.takvim_widget.setMinimumHeight(350)
        self.takvim_widget.clicked.connect(self.takvim_tarih_secildi)
        sol_layout.addWidget(self.takvim_widget)
        
        # Seçili tarihteki etkinlikler
        self.secili_tarih_label = QLabel("Bugun")
        self.secili_tarih_label.setFont(QFont("Arial", 11, QFont.Bold))
        sol_layout.addWidget(self.secili_tarih_label)
        
        self.gunluk_etkinlik_tablo = QTableWidget()
        self.gunluk_etkinlik_tablo.setColumnCount(3)
        self.gunluk_etkinlik_tablo.setHorizontalHeaderLabels(["Saat", "Baslik", "Tip"])
        self.gunluk_etkinlik_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.gunluk_etkinlik_tablo.setMaximumHeight(200)
        
        sol_layout.addWidget(self.gunluk_etkinlik_tablo)
        sol_widget.setLayout(sol_layout)
        
        # SAG: TUM ETKINLIK LISTESI
        sag_widget = QWidget()
        sag_layout = QVBoxLayout()
        sag_layout.setContentsMargins(0, 0, 0, 0)
        
        sag_baslik_layout = QHBoxLayout()
        sag_baslik = QLabel("Tum Etkinlikler")
        sag_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        sag_baslik_layout.addWidget(sag_baslik)
        sag_baslik_layout.addStretch()
        
        # Filtre
        sag_baslik_layout.addWidget(QLabel("Filtre:"))
        self.etkinlik_filtre = QComboBox()
        self.etkinlik_filtre.addItems(["Hepsi", "Bugun", "Bu Hafta", "Bu Ay", "Yaklaşan", "Gecmis", "Planlandi", "Tamamlandi"])
        self.etkinlik_filtre.currentTextChanged.connect(self.etkinlik_yenile)
        sag_baslik_layout.addWidget(self.etkinlik_filtre)
        
        sag_layout.addLayout(sag_baslik_layout)
        
        self.etkinlik_tablo = QTableWidget()
        self.etkinlik_tablo.setColumnCount(8)
        self.etkinlik_tablo.setHorizontalHeaderLabels([
            "ID", "Tarih", "Saat", "Baslik", "Tip", "Musteri", "Oncelik", "Durum"
        ])
        self.etkinlik_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        sag_layout.addWidget(self.etkinlik_tablo)
        sag_widget.setLayout(sag_layout)
        
        ana_layout.addWidget(sol_widget)
        ana_layout.addWidget(sag_widget)
        
        layout.addLayout(ana_layout)
        
        widget.setLayout(layout)
        return widget
    
    def takvim_tarih_secildi(self, qdate):
        tarih = qdate.toString("yyyy-MM-dd")
        self.secili_tarih_label.setText(f"Tarih: {qdate.toString('dd MMMM yyyy')}")
        
        etkinlikler = self.db.etkinlikleri_getir()
        gun_etkinlikleri = [e for e in etkinlikler if e['tarih'] == tarih]
        
        self.gunluk_etkinlik_tablo.setRowCount(0)
        for e in gun_etkinlikleri:
            row = self.gunluk_etkinlik_tablo.rowCount()
            self.gunluk_etkinlik_tablo.insertRow(row)
            
            self.gunluk_etkinlik_tablo.setItem(row, 0, QTableWidgetItem(e['saat'] or ""))
            self.gunluk_etkinlik_tablo.setItem(row, 1, QTableWidgetItem(e['baslik']))
            
            tip_item = QTableWidgetItem(e['tip'])
            if e['durum'] == 'Tamamlandi':
                tip_item.setForeground(QColor(COLORS['success']))
            elif e['durum'] == 'Iptal':
                tip_item.setForeground(QColor(COLORS['danger']))
            else:
                tip_item.setForeground(QColor(COLORS['primary_light']))
            self.gunluk_etkinlik_tablo.setItem(row, 2, tip_item)
    
    def etkinlik_ekle(self):
        musteriler = self.db.musterileri_getir()
        dialog = EtkinlikDialog(self, musteriler)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.etkinlik_ekle(
                veri['musteri_id'], veri['baslik'], veri['aciklama'], veri['tip'],
                veri['oncelik'], veri['tarih'], veri['saat'], veri['sure'],
                veri['konum'], veri['durum'], veri['hatirlatici']
            )
            QMessageBox.information(self, "Basarili", "Etkinlik eklendi!")
            self.veriyi_yenile()
    
    def etkinlik_duzenle(self):
        row = self.etkinlik_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Etkinlik secin!")
            return
        
        etkinlik_id = int(self.etkinlik_tablo.item(row, 0).text())
        etkinlikler = self.db.etkinlikleri_getir()
        etkinlik = next((e for e in etkinlikler if e['id'] == etkinlik_id), None)
        musteriler = self.db.musterileri_getir()
        
        dialog = EtkinlikDialog(self, musteriler, etkinlik)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.etkinlik_guncelle(
                etkinlik_id, veri['musteri_id'], veri['baslik'], veri['aciklama'], veri['tip'],
                veri['oncelik'], veri['tarih'], veri['saat'], veri['sure'],
                veri['konum'], veri['durum'], veri['hatirlatici']
            )
            QMessageBox.information(self, "Basarili", "Etkinlik guncellendi!")
            self.veriyi_yenile()
    
    def etkinlik_sil(self):
        row = self.etkinlik_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Etkinlik secin!")
            return
        
        etkinlik_id = int(self.etkinlik_tablo.item(row, 0).text())
        baslik = self.etkinlik_tablo.item(row, 3).text()
        
        cevap = QMessageBox.question(self, "Silme Onay", f"{baslik} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.etkinlik_sil(etkinlik_id)
            QMessageBox.information(self, "Basarili", "Etkinlik silindi!")
            self.veriyi_yenile()
    
    def etkinlik_durum_degistir(self, yeni_durum):
        row = self.etkinlik_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Etkinlik secin!")
            return
        
        etkinlik_id = int(self.etkinlik_tablo.item(row, 0).text())
        self.db.etkinlik_durum_guncelle(etkinlik_id, yeni_durum)
        QMessageBox.information(self, "Basarili", f"Etkinlik '{yeni_durum}' olarak isaretlendi!")
        self.veriyi_yenile()
    
    def etkinlik_yenile(self):
        if not hasattr(self, 'etkinlik_tablo'):
            return
        
        etkinlikler = self.db.etkinlikleri_getir()
        bugun_str = datetime.now().strftime('%Y-%m-%d')
        
        # KPI Hesapla
        bugun_etkinlikleri = [e for e in etkinlikler if e['tarih'] == bugun_str]
        
        hafta_basla = datetime.now().date()
        hafta_bitis = hafta_basla + timedelta(days=7)
        hafta_etkinlikleri = [e for e in etkinlikler if e['tarih'] and hafta_basla.strftime('%Y-%m-%d') <= e['tarih'] <= hafta_bitis.strftime('%Y-%m-%d')]
        
        yaklasan = [e for e in etkinlikler if e['durum'] == 'Planlandi' and e['tarih'] >= bugun_str and e['oncelik'] in ['Yuksek', 'Acil']]
        tamamlanan = [e for e in etkinlikler if e['durum'] == 'Tamamlandi']
        
        self.kpi_bugun_etkinlik.deger_label.setText(str(len(bugun_etkinlikleri)))
        self.kpi_bu_hafta.deger_label.setText(str(len(hafta_etkinlikleri)))
        self.kpi_yaklasan.deger_label.setText(str(len(yaklasan)))
        self.kpi_tamamlanan.deger_label.setText(str(len(tamamlanan)))
        
        # Filtre uygula
        filtre = self.etkinlik_filtre.currentText() if hasattr(self, 'etkinlik_filtre') else "Hepsi"
        
        if filtre == "Bugun":
            filtrelenmis = bugun_etkinlikleri
        elif filtre == "Bu Hafta":
            filtrelenmis = hafta_etkinlikleri
        elif filtre == "Bu Ay":
            ay_basla = datetime.now().replace(day=1).strftime('%Y-%m-%d')
            ay_bitis = (datetime.now().replace(day=1) + timedelta(days=31)).strftime('%Y-%m-%d')
            filtrelenmis = [e for e in etkinlikler if e['tarih'] and ay_basla <= e['tarih'] <= ay_bitis]
        elif filtre == "Yaklaşan":
            filtrelenmis = [e for e in etkinlikler if e['tarih'] >= bugun_str and e['durum'] == 'Planlandi']
        elif filtre == "Gecmis":
            filtrelenmis = [e for e in etkinlikler if e['tarih'] < bugun_str]
        elif filtre in ["Planlandi", "Tamamlandi"]:
            filtrelenmis = [e for e in etkinlikler if e['durum'] == filtre]
        else:
            filtrelenmis = etkinlikler
        
        self.etkinlik_tablo.setRowCount(0)
        for e in filtrelenmis:
            row = self.etkinlik_tablo.rowCount()
            self.etkinlik_tablo.insertRow(row)
            
            self.etkinlik_tablo.setItem(row, 0, QTableWidgetItem(str(e['id'])))
            
            tarih_item = QTableWidgetItem(e['tarih'])
            if e['tarih'] == bugun_str:
                tarih_item.setForeground(QColor(COLORS['warning']))
            elif e['tarih'] < bugun_str:
                tarih_item.setForeground(QColor(COLORS['text_sec']))
            self.etkinlik_tablo.setItem(row, 1, tarih_item)
            
            self.etkinlik_tablo.setItem(row, 2, QTableWidgetItem(e['saat'] or ""))
            self.etkinlik_tablo.setItem(row, 3, QTableWidgetItem(e['baslik']))
            self.etkinlik_tablo.setItem(row, 4, QTableWidgetItem(e['tip']))
            self.etkinlik_tablo.setItem(row, 5, QTableWidgetItem(e['musteri_adi'] or "-"))
            
            oncelik_item = QTableWidgetItem(e['oncelik'])
            if e['oncelik'] == 'Acil':
                oncelik_item.setForeground(QColor(COLORS['danger']))
            elif e['oncelik'] == 'Yuksek':
                oncelik_item.setForeground(QColor(COLORS['warning']))
            elif e['oncelik'] == 'Orta':
                oncelik_item.setForeground(QColor(COLORS['primary_light']))
            else:
                oncelik_item.setForeground(QColor(COLORS['text_sec']))
            self.etkinlik_tablo.setItem(row, 6, oncelik_item)
            
            durum_item = QTableWidgetItem(e['durum'])
            if e['durum'] == 'Tamamlandi':
                durum_item.setForeground(QColor(COLORS['success']))
            elif e['durum'] == 'Iptal':
                durum_item.setForeground(QColor(COLORS['danger']))
            elif e['durum'] == 'Ertelendi':
                durum_item.setForeground(QColor(COLORS['warning']))
            else:
                durum_item.setForeground(QColor(COLORS['primary']))
            self.etkinlik_tablo.setItem(row, 7, durum_item)
        
        # Takvimde bugünü göster
        if hasattr(self, 'takvim_widget'):
            self.takvim_tarih_secildi(self.takvim_widget.selectedDate())
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 8: FIRSAT / PIPELINE YONETIMI
    # ─────────────────────────────────────────────────────────────────────────
    
    def firsat_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # KPI Kartlari
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.kpi_aktif_firsat = self.kpi_kart("Aktif Firsat", "0", COLORS['primary'])
        self.kpi_pipeline_deger = self.kpi_kart("Pipeline Degeri", "0 TRY", COLORS['warning'])
        self.kpi_agirlikli = self.kpi_kart("Agirlikli Tahmin", "0 TRY", COLORS['success'])
        self.kpi_kazanilan = self.kpi_kart("Kazanma Orani", "0%", COLORS['danger'])
        
        kpi_layout.addWidget(self.kpi_aktif_firsat)
        kpi_layout.addWidget(self.kpi_pipeline_deger)
        kpi_layout.addWidget(self.kpi_agirlikli)
        kpi_layout.addWidget(self.kpi_kazanilan)
        
        layout.addLayout(kpi_layout)
        
        # Aksiyon Butonlari
        aksiyon_layout = QHBoxLayout()
        
        ekle_btn = QPushButton("Yeni Firsat")
        ekle_btn.clicked.connect(self.firsat_ekle)
        
        duzenle_btn = QPushButton("Duzenle")
        duzenle_btn.clicked.connect(self.firsat_duzenle)
        
        sil_btn = QPushButton("Sil")
        sil_btn.clicked.connect(self.firsat_sil)
        
        ilerlet_btn = QPushButton("Asamayi Ilerlet")
        ilerlet_btn.clicked.connect(self.firsat_ilerlet)
        
        kazanildi_btn = QPushButton("Kazanildi")
        kazanildi_btn.clicked.connect(lambda: self.firsat_asama_degistir("Kazanildi"))
        
        kaybedildi_btn = QPushButton("Kaybedildi")
        kaybedildi_btn.clicked.connect(lambda: self.firsat_asama_degistir("Kaybedildi"))
        
        aksiyon_layout.addWidget(ekle_btn)
        aksiyon_layout.addWidget(duzenle_btn)
        aksiyon_layout.addWidget(sil_btn)
        aksiyon_layout.addStretch()
        aksiyon_layout.addWidget(ilerlet_btn)
        aksiyon_layout.addWidget(kazanildi_btn)
        aksiyon_layout.addWidget(kaybedildi_btn)
        
        layout.addLayout(aksiyon_layout)
        
        # Filtre
        filtre_layout = QHBoxLayout()
        filtre_layout.addWidget(QLabel("Asama Filtre:"))
        self.firsat_filtre = QComboBox()
        self.firsat_filtre.addItems(["Hepsi", "Yeni", "Iletisim", "Teklif", "Pazarlik", "Kazanildi", "Kaybedildi"])
        self.firsat_filtre.currentTextChanged.connect(self.firsat_yenile)
        filtre_layout.addWidget(self.firsat_filtre)
        filtre_layout.addStretch()
        
        layout.addLayout(filtre_layout)
        
        # Firsat Tablosu
        self.firsat_tablo = QTableWidget()
        self.firsat_tablo.setColumnCount(9)
        self.firsat_tablo.setHorizontalHeaderLabels([
            "ID", "Musteri", "Baslik", "Asama", "Deger", "Olasilik", "Agirlikli", "Son Iletisim", "Kapanis"
        ])
        self.firsat_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.firsat_tablo)
        
        # Pipeline Grafigi
        grafik_layout = QHBoxLayout()
        grafik_layout.setSpacing(15)
        
        self.pipeline_grafik = self.grafik_widget()
        pipeline_frame = QFrame()
        pipeline_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        pipeline_vlayout = QVBoxLayout()
        pipeline_baslik = QLabel("Pipeline Asamalari")
        pipeline_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        pipeline_vlayout.addWidget(pipeline_baslik)
        pipeline_vlayout.addWidget(self.pipeline_grafik)
        pipeline_frame.setLayout(pipeline_vlayout)
        
        self.firsat_deger_grafik = self.grafik_widget()
        deger_frame = QFrame()
        deger_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        deger_vlayout = QVBoxLayout()
        deger_baslik = QLabel("Asama Bazli Deger")
        deger_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        deger_vlayout.addWidget(deger_baslik)
        deger_vlayout.addWidget(self.firsat_deger_grafik)
        deger_frame.setLayout(deger_vlayout)
        
        grafik_layout.addWidget(pipeline_frame)
        grafik_layout.addWidget(deger_frame)
        
        layout.addLayout(grafik_layout)
        
        widget.setLayout(layout)
        return widget
    
    def firsat_ekle(self):
        musteriler = self.db.musterileri_getir()
        if not musteriler:
            QMessageBox.warning(self, "Uyari", "Once musteri ekleyin!")
            return
        
        dialog = FirsatDialog(self, musteriler)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.firsat_ekle(
                veri['musteri_id'], veri['baslik'], veri['aciklama'], veri['deger'],
                veri['olasilik'], veri['asama'], veri['son_iletisim'], veri['beklenen'], veri['notlar']
            )
            QMessageBox.information(self, "Basarili", "Firsat eklendi!")
            self.veriyi_yenile()
    
    def firsat_duzenle(self):
        row = self.firsat_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Firsat secin!")
            return
        
        firsat_id = int(self.firsat_tablo.item(row, 0).text())
        firsatlar = self.db.firsatlari_getir()
        firsat = next((f for f in firsatlar if f['id'] == firsat_id), None)
        musteriler = self.db.musterileri_getir()
        
        dialog = FirsatDialog(self, musteriler, firsat)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.firsat_guncelle(
                firsat_id, veri['baslik'], veri['aciklama'], veri['deger'],
                veri['olasilik'], veri['asama'], veri['son_iletisim'], veri['beklenen'], veri['notlar']
            )
            QMessageBox.information(self, "Basarili", "Firsat guncellendi!")
            self.veriyi_yenile()
    
    def firsat_sil(self):
        row = self.firsat_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Firsat secin!")
            return
        
        firsat_id = int(self.firsat_tablo.item(row, 0).text())
        baslik = self.firsat_tablo.item(row, 2).text()
        
        cevap = QMessageBox.question(self, "Silme Onay", f"{baslik} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.firsat_sil(firsat_id)
            QMessageBox.information(self, "Basarili", "Firsat silindi!")
            self.veriyi_yenile()
    
    def firsat_ilerlet(self):
        row = self.firsat_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Firsat secin!")
            return
        
        firsat_id = int(self.firsat_tablo.item(row, 0).text())
        firsatlar = self.db.firsatlari_getir()
        firsat = next((f for f in firsatlar if f['id'] == firsat_id), None)
        
        asama_sira = ["Yeni", "Iletisim", "Teklif", "Pazarlik", "Kazanildi"]
        
        if firsat['asama'] in ["Kazanildi", "Kaybedildi"]:
            QMessageBox.warning(self, "Uyari", "Firsat zaten kapali!")
            return
        
        mevcut_index = asama_sira.index(firsat['asama'])
        if mevcut_index < len(asama_sira) - 1:
            yeni_asama = asama_sira[mevcut_index + 1]
            self.db.firsat_asama_guncelle(firsat_id, yeni_asama)
            QMessageBox.information(self, "Basarili", f"Asama '{yeni_asama}' olarak guncellendi!")
            self.veriyi_yenile()
    
    def firsat_asama_degistir(self, yeni_asama):
        row = self.firsat_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Firsat secin!")
            return
        
        firsat_id = int(self.firsat_tablo.item(row, 0).text())
        cevap = QMessageBox.question(self, "Onay", f"Firsat '{yeni_asama}' olarak isaretlensin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.firsat_asama_guncelle(firsat_id, yeni_asama)
            QMessageBox.information(self, "Basarili", f"Firsat '{yeni_asama}' olarak guncellendi!")
            self.veriyi_yenile()
    
    def firsat_yenile(self):
        if not hasattr(self, 'firsat_tablo'):
            return
        
        firsatlar = self.db.firsatlari_getir()
        
        # Filtre uygula
        filtre = self.firsat_filtre.currentText() if hasattr(self, 'firsat_filtre') else "Hepsi"
        if filtre != "Hepsi":
            filtrelenmis = [f for f in firsatlar if f['asama'] == filtre]
        else:
            filtrelenmis = firsatlar
        
        # KPI Hesapla (tum firsatlar uzerinden)
        aktif = [f for f in firsatlar if f['asama'] not in ["Kazanildi", "Kaybedildi"]]
        kazanildi = [f for f in firsatlar if f['asama'] == "Kazanildi"]
        kapali = [f for f in firsatlar if f['asama'] in ["Kazanildi", "Kaybedildi"]]
        
        pipeline_deger = sum(f['tahmini_deger'] for f in aktif)
        agirlikli = sum(f['tahmini_deger'] * f['olasilik'] / 100 for f in aktif)
        
        kazanma_orani = (len(kazanildi) / len(kapali) * 100) if kapali else 0
        
        self.kpi_aktif_firsat.deger_label.setText(str(len(aktif)))
        self.kpi_pipeline_deger.deger_label.setText(f"{pipeline_deger:,.0f} TRY")
        self.kpi_agirlikli.deger_label.setText(f"{agirlikli:,.0f} TRY")
        self.kpi_kazanilan.deger_label.setText(f"{kazanma_orani:.1f}%")
        
        # Tablo
        self.firsat_tablo.setRowCount(0)
        for f in filtrelenmis:
            row = self.firsat_tablo.rowCount()
            self.firsat_tablo.insertRow(row)
            
            self.firsat_tablo.setItem(row, 0, QTableWidgetItem(str(f['id'])))
            self.firsat_tablo.setItem(row, 1, QTableWidgetItem(f['musteri_adi']))
            self.firsat_tablo.setItem(row, 2, QTableWidgetItem(f['baslik']))
            
            asama_item = QTableWidgetItem(f['asama'])
            if f['asama'] == 'Kazanildi':
                asama_item.setForeground(QColor(COLORS['success']))
            elif f['asama'] == 'Kaybedildi':
                asama_item.setForeground(QColor(COLORS['danger']))
            elif f['asama'] == 'Pazarlik':
                asama_item.setForeground(QColor(COLORS['warning']))
            elif f['asama'] == 'Teklif':
                asama_item.setForeground(QColor(COLORS['primary_light']))
            else:
                asama_item.setForeground(QColor(COLORS['secondary']))
            self.firsat_tablo.setItem(row, 3, asama_item)
            
            self.firsat_tablo.setItem(row, 4, QTableWidgetItem(f"{f['tahmini_deger']:,.0f} TRY"))
            
            olasilik_item = QTableWidgetItem(f"%{f['olasilik']}")
            if f['olasilik'] >= 75:
                olasilik_item.setForeground(QColor(COLORS['success']))
            elif f['olasilik'] >= 50:
                olasilik_item.setForeground(QColor(COLORS['warning']))
            else:
                olasilik_item.setForeground(QColor(COLORS['danger']))
            self.firsat_tablo.setItem(row, 5, olasilik_item)
            
            agirlikli_deger = f['tahmini_deger'] * f['olasilik'] / 100
            self.firsat_tablo.setItem(row, 6, QTableWidgetItem(f"{agirlikli_deger:,.0f} TRY"))
            
            self.firsat_tablo.setItem(row, 7, QTableWidgetItem(f['son_iletisim'] or ""))
            self.firsat_tablo.setItem(row, 8, QTableWidgetItem(f['beklenen_kapanis'] or ""))
        
        # Pipeline Grafik (Bar - Funnel benzeri)
        self.pipeline_grafik.fig.clear()
        ax = self.pipeline_grafik.fig.add_subplot(111)
        
        asama_sira = ["Yeni", "Iletisim", "Teklif", "Pazarlik", "Kazanildi", "Kaybedildi"]
        asama_say = {a: 0 for a in asama_sira}
        for f in firsatlar:
            if f['asama'] in asama_say:
                asama_say[f['asama']] += 1
        
        asamalar = list(asama_say.keys())
        sayilar = list(asama_say.values())
        
        renkler = [COLORS['secondary'], COLORS['primary_light'], COLORS['primary'], 
                   COLORS['warning'], COLORS['success'], COLORS['danger']]
        
        ax.barh(asamalar, sayilar, color=renkler, alpha=0.8)
        ax.invert_yaxis()
        ax.set_xlabel('Firsat Sayisi', color=COLORS['text_sec'])
        ax.tick_params(colors=COLORS['text_sec'])
        ax.set_facecolor(COLORS['bg_secondary'])
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.spines['left'].set_color(COLORS['border'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.1, axis='x')
        
        # Sayıları çubukların üstüne yaz
        for i, v in enumerate(sayilar):
            if v > 0:
                ax.text(v, i, f' {v}', color=COLORS['text_main'], va='center', fontsize=10)
        
        self.pipeline_grafik.fig.tight_layout()
        self.pipeline_grafik.draw()
        
        # Asama Bazli Deger Grafik
        self.firsat_deger_grafik.fig.clear()
        ax2 = self.firsat_deger_grafik.fig.add_subplot(111)
        
        asama_deger = {a: 0 for a in asama_sira}
        for f in firsatlar:
            if f['asama'] in asama_deger:
                asama_deger[f['asama']] += f['tahmini_deger']
        
        degerler = list(asama_deger.values())
        ax2.bar(asamalar, degerler, color=renkler, alpha=0.8)
        ax2.set_ylabel('Toplam Deger (TRY)', color=COLORS['text_sec'])
        ax2.tick_params(colors=COLORS['text_sec'])
        ax2.tick_params(axis='x', rotation=45)
        ax2.set_facecolor(COLORS['bg_secondary'])
        ax2.spines['bottom'].set_color(COLORS['border'])
        ax2.spines['left'].set_color(COLORS['border'])
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
        ax2.grid(True, alpha=0.1, axis='y')
        
        self.firsat_deger_grafik.fig.tight_layout()
        self.firsat_deger_grafik.draw()
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 7: PAZARLAMA KAMPANYALARI
    # ─────────────────────────────────────────────────────────────────────────
    
    def kampanya_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # KPI Kartlari
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.kpi_aktif_kampanya = self.kpi_kart("Aktif Kampanya", "0", COLORS['success'])
        self.kpi_toplam_katilim = self.kpi_kart("Toplam Katilim", "0", COLORS['primary'])
        self.kpi_donusum = self.kpi_kart("Ort. Donusum", "0%", COLORS['warning'])
        self.kpi_butce = self.kpi_kart("Toplam Butce", "0 TRY", COLORS['danger'])
        
        kpi_layout.addWidget(self.kpi_aktif_kampanya)
        kpi_layout.addWidget(self.kpi_toplam_katilim)
        kpi_layout.addWidget(self.kpi_donusum)
        kpi_layout.addWidget(self.kpi_butce)
        
        layout.addLayout(kpi_layout)
        
        # Aksiyon Butonlari
        aksiyon_layout = QHBoxLayout()
        
        ekle_btn = QPushButton("Yeni Kampanya")
        ekle_btn.clicked.connect(self.kampanya_ekle)
        
        duzenle_btn = QPushButton("Duzenle")
        duzenle_btn.clicked.connect(self.kampanya_duzenle)
        
        sil_btn = QPushButton("Sil")
        sil_btn.clicked.connect(self.kampanya_sil)
        
        katilimci_ekle_btn = QPushButton("Katilimci Ekle")
        katilimci_ekle_btn.clicked.connect(self.kampanyaya_katilimci_ekle)
        
        detay_btn = QPushButton("Detay")
        detay_btn.clicked.connect(self.kampanya_detay_goster)
        
        aksiyon_layout.addWidget(ekle_btn)
        aksiyon_layout.addWidget(duzenle_btn)
        aksiyon_layout.addWidget(sil_btn)
        aksiyon_layout.addStretch()
        aksiyon_layout.addWidget(katilimci_ekle_btn)
        aksiyon_layout.addWidget(detay_btn)
        
        layout.addLayout(aksiyon_layout)
        
        # Kampanya Tablosu
        layout.addWidget(QLabel("Kampanyalar"))
        
        self.kampanya_tablo = QTableWidget()
        self.kampanya_tablo.setColumnCount(9)
        self.kampanya_tablo.setHorizontalHeaderLabels([
            "ID", "Ad", "Tip", "Indirim", "Butce", "Segment", "Baslama", "Bitis", "Durum"
        ])
        self.kampanya_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.kampanya_tablo)
        
        # Performans Grafigi
        grafik_layout = QHBoxLayout()
        grafik_layout.setSpacing(15)
        
        self.kampanya_grafik = self.grafik_widget()
        kampanya_frame = QFrame()
        kampanya_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        kampanya_vlayout = QVBoxLayout()
        kampanya_baslik = QLabel("Kampanya Performansi")
        kampanya_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        kampanya_vlayout.addWidget(kampanya_baslik)
        kampanya_vlayout.addWidget(self.kampanya_grafik)
        kampanya_frame.setLayout(kampanya_vlayout)
        
        self.segment_grafik = self.grafik_widget()
        segment_frame = QFrame()
        segment_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        segment_vlayout = QVBoxLayout()
        segment_baslik = QLabel("Segment Dagilimi")
        segment_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        segment_vlayout.addWidget(segment_baslik)
        segment_vlayout.addWidget(self.segment_grafik)
        segment_frame.setLayout(segment_vlayout)
        
        grafik_layout.addWidget(kampanya_frame)
        grafik_layout.addWidget(segment_frame)
        
        layout.addLayout(grafik_layout)
        
        widget.setLayout(layout)
        return widget
    
    def kampanya_ekle(self):
        dialog = KampanyaDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.kampanya_ekle(
                veri['ad'], veri['aciklama'], veri['tip'], veri['indirim'],
                veri['butce'], veri['segment'], veri['basla'], veri['bitis'], veri['durum']
            )
            QMessageBox.information(self, "Basarili", "Kampanya eklendi!")
            self.veriyi_yenile()
    
    def kampanya_duzenle(self):
        row = self.kampanya_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Kampanya secin!")
            return
        
        kampanya_id = int(self.kampanya_tablo.item(row, 0).text())
        kampanyalar = self.db.kampanyalari_getir()
        kampanya = next((k for k in kampanyalar if k['id'] == kampanya_id), None)
        
        dialog = KampanyaDialog(self, kampanya)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.kampanya_guncelle(
                kampanya_id, veri['ad'], veri['aciklama'], veri['tip'], veri['indirim'],
                veri['butce'], veri['segment'], veri['basla'], veri['bitis'], veri['durum']
            )
            QMessageBox.information(self, "Basarili", "Kampanya guncellendi!")
            self.veriyi_yenile()
    
    def kampanya_sil(self):
        row = self.kampanya_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Kampanya secin!")
            return
        
        kampanya_id = int(self.kampanya_tablo.item(row, 0).text())
        ad = self.kampanya_tablo.item(row, 1).text()
        
        cevap = QMessageBox.question(self, "Silme Onay", f"{ad} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.kampanya_sil(kampanya_id)
            QMessageBox.information(self, "Basarili", "Kampanya silindi!")
            self.veriyi_yenile()
    
    def kampanyaya_katilimci_ekle(self):
        row = self.kampanya_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Kampanya secin!")
            return
        
        kampanya_id = int(self.kampanya_tablo.item(row, 0).text())
        kampanya_ad = self.kampanya_tablo.item(row, 1).text()
        
        musteriler = self.db.musterileri_getir()
        if not musteriler:
            QMessageBox.warning(self, "Uyari", "Musteri yok!")
            return
        
        # Tüm uygun müşterileri ekle
        kampanyalar = self.db.kampanyalari_getir()
        kampanya = next((k for k in kampanyalar if k['id'] == kampanya_id), None)
        segment = kampanya['hedef_segment']
        
        if segment == "Aktif Musteriler":
            hedef = [m for m in musteriler if m['durum'] == 'Aktif']
        elif segment == "Pasif Musteriler":
            hedef = [m for m in musteriler if m['durum'] == 'Pasif']
        elif segment == "VIP Musteriler":
            hedef = [m for m in musteriler if m['toplam_harcama'] > 10000]
        else:
            hedef = musteriler
        
        for m in hedef:
            try:
                self.db.kampanya_katilim_ekle(kampanya_id, m['id'])
            except:
                pass
        
        QMessageBox.information(self, "Basarili", f"{len(hedef)} musteri '{kampanya_ad}' kampanyasina eklendi!")
        self.veriyi_yenile()
    
    def kampanya_detay_goster(self):
        row = self.kampanya_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Kampanya secin!")
            return
        
        kampanya_id = int(self.kampanya_tablo.item(row, 0).text())
        kampanya_ad = self.kampanya_tablo.item(row, 1).text()
        
        istatistik = self.db.kampanya_istatistik(kampanya_id)
        katilimcilar = self.db.kampanya_katilimcilari_getir(kampanya_id)
        
        mesaj = f"""KAMPANYA: {kampanya_ad}

Toplam Katilimci: {istatistik['toplam_katilim']}
Kullanan: {istatistik['kullanan']}
Donusum Orani: {istatistik['donusum_orani']:.1f}%

Son Katilimcilar:
"""
        for k in katilimcilar[:10]:
            durum = "✓" if k['kullanildi'] else "○"
            mesaj += f"\n{durum} {k['musteri_adi']} - {k['tarih']}"
        
        QMessageBox.information(self, "Kampanya Detay", mesaj)
    
    def kampanya_yenile(self):
        if not hasattr(self, 'kampanya_tablo'):
            return
        
        kampanyalar = self.db.kampanyalari_getir()
        
        # KPI
        aktif = sum(1 for k in kampanyalar if k['durum'] == 'Aktif')
        toplam_butce = sum(k['butce'] for k in kampanyalar)
        
        toplam_katilim = 0
        toplam_kullanan = 0
        for k in kampanyalar:
            istatistik = self.db.kampanya_istatistik(k['id'])
            toplam_katilim += istatistik['toplam_katilim']
            toplam_kullanan += istatistik['kullanan']
        
        ort_donusum = (toplam_kullanan / toplam_katilim * 100) if toplam_katilim > 0 else 0
        
        self.kpi_aktif_kampanya.deger_label.setText(str(aktif))
        self.kpi_toplam_katilim.deger_label.setText(str(toplam_katilim))
        self.kpi_donusum.deger_label.setText(f"{ort_donusum:.1f}%")
        self.kpi_butce.deger_label.setText(f"{toplam_butce:,.0f} TRY")
        
        # Kampanya Tablo
        self.kampanya_tablo.setRowCount(0)
        for k in kampanyalar:
            row = self.kampanya_tablo.rowCount()
            self.kampanya_tablo.insertRow(row)
            
            self.kampanya_tablo.setItem(row, 0, QTableWidgetItem(str(k['id'])))
            self.kampanya_tablo.setItem(row, 1, QTableWidgetItem(k['ad']))
            self.kampanya_tablo.setItem(row, 2, QTableWidgetItem(k['tip']))
            self.kampanya_tablo.setItem(row, 3, QTableWidgetItem(f"%{k['indirim_orani']:.0f}"))
            self.kampanya_tablo.setItem(row, 4, QTableWidgetItem(f"{k['butce']:,.0f} TRY"))
            self.kampanya_tablo.setItem(row, 5, QTableWidgetItem(k['hedef_segment']))
            self.kampanya_tablo.setItem(row, 6, QTableWidgetItem(k['basla_tarih']))
            self.kampanya_tablo.setItem(row, 7, QTableWidgetItem(k['bitis_tarih']))
            
            durum_item = QTableWidgetItem(k['durum'])
            if k['durum'] == 'Aktif':
                durum_item.setForeground(QColor(COLORS['success']))
            elif k['durum'] == 'Tamamlandi':
                durum_item.setForeground(QColor(COLORS['primary']))
            elif k['durum'] == 'Duraklatildi':
                durum_item.setForeground(QColor(COLORS['warning']))
            elif k['durum'] == 'Iptal':
                durum_item.setForeground(QColor(COLORS['danger']))
            else:
                durum_item.setForeground(QColor(COLORS['text_sec']))
            self.kampanya_tablo.setItem(row, 8, durum_item)
        
        # Kampanya Performans Grafigi
        self.kampanya_grafik.fig.clear()
        ax = self.kampanya_grafik.fig.add_subplot(111)
        
        aktif_kampanyalar = [k for k in kampanyalar if k['durum'] in ['Aktif', 'Tamamlandi']][:8]
        if aktif_kampanyalar:
            adlar = [k['ad'][:15] for k in aktif_kampanyalar]
            katilimlar = []
            kullananlar = []
            
            for k in aktif_kampanyalar:
                ist = self.db.kampanya_istatistik(k['id'])
                katilimlar.append(ist['toplam_katilim'])
                kullananlar.append(ist['kullanan'])
            
            x = range(len(adlar))
            width = 0.35
            ax.bar([i - width/2 for i in x], katilimlar, width, color=COLORS['primary'], alpha=0.8, label='Katilim')
            ax.bar([i + width/2 for i in x], kullananlar, width, color=COLORS['success'], alpha=0.8, label='Kullanan')
            
            ax.set_xticks(x)
            ax.set_xticklabels(adlar, color=COLORS['text_sec'], rotation=45, ha='right', fontsize=8)
            ax.tick_params(colors=COLORS['text_sec'])
            ax.legend(facecolor=COLORS['bg_secondary'], edgecolor=COLORS['border'], labelcolor=COLORS['text_main'])
        
        ax.set_facecolor(COLORS['bg_secondary'])
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.spines['left'].set_color(COLORS['border'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.1)
        
        self.kampanya_grafik.fig.tight_layout()
        self.kampanya_grafik.draw()
        
        # Segment Pasta Grafigi
        self.segment_grafik.fig.clear()
        ax2 = self.segment_grafik.fig.add_subplot(111)
        
        segment_say = {}
        for k in kampanyalar:
            seg = k['hedef_segment']
            segment_say[seg] = segment_say.get(seg, 0) + 1
        
        if segment_say:
            labels = list(segment_say.keys())
            sizes = list(segment_say.values())
            colors_pie = [COLORS['primary'], COLORS['success'], COLORS['warning'], COLORS['danger'], COLORS['secondary']]
            ax2.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors_pie[:len(labels)], startangle=90)
        
        for text in ax2.texts:
            text.set_color(COLORS['text_main'])
            text.set_fontsize(9)
        
        self.segment_grafik.draw()
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 6: EMAIL OTOMASYONU
    # ─────────────────────────────────────────────────────────────────────────
    
    def email_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # KPI Kartlari
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.kpi_sablon_sayi = self.kpi_kart("Sablon Sayisi", "0", COLORS['primary'])
        self.kpi_gonderilen = self.kpi_kart("Gonderilen Email", "0", COLORS['success'])
        self.kpi_bugun_gonderilen = self.kpi_kart("Bugun Gonderilen", "0", COLORS['warning'])
        
        kpi_layout.addWidget(self.kpi_sablon_sayi)
        kpi_layout.addWidget(self.kpi_gonderilen)
        kpi_layout.addWidget(self.kpi_bugun_gonderilen)
        
        layout.addLayout(kpi_layout)
        
        # Aksiyon Butonlari
        aksiyon_layout = QHBoxLayout()
        
        sablon_ekle_btn = QPushButton("Yeni Sablon")
        sablon_ekle_btn.clicked.connect(self.email_sablon_ekle)
        
        sablon_duzenle_btn = QPushButton("Sablon Duzenle")
        sablon_duzenle_btn.clicked.connect(self.email_sablon_duzenle)
        
        sablon_sil_btn = QPushButton("Sablon Sil")
        sablon_sil_btn.clicked.connect(self.email_sablon_sil)
        
        email_gonder_btn = QPushButton("Email Gonder")
        email_gonder_btn.clicked.connect(self.email_gonder)
        
        aksiyon_layout.addWidget(sablon_ekle_btn)
        aksiyon_layout.addWidget(sablon_duzenle_btn)
        aksiyon_layout.addWidget(sablon_sil_btn)
        aksiyon_layout.addStretch()
        aksiyon_layout.addWidget(email_gonder_btn)
        
        layout.addLayout(aksiyon_layout)
        
        # Sablon Tablosu
        layout.addWidget(QLabel("Email Sablonlari"))
        
        self.sablon_tablo = QTableWidget()
        self.sablon_tablo.setColumnCount(5)
        self.sablon_tablo.setHorizontalHeaderLabels(["ID", "Ad", "Kategori", "Konu", "Onizleme"])
        self.sablon_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sablon_tablo.setMaximumHeight(250)
        
        layout.addWidget(self.sablon_tablo)
        
        # Email Gecmis Tablosu
        layout.addWidget(QLabel("Gonderilen Emailler"))
        
        self.email_gecmis_tablo = QTableWidget()
        self.email_gecmis_tablo.setColumnCount(5)
        self.email_gecmis_tablo.setHorizontalHeaderLabels(["ID", "Musteri", "Email", "Konu", "Tarih"])
        self.email_gecmis_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.email_gecmis_tablo)
        
        widget.setLayout(layout)
        return widget
    
    def email_sablon_ekle(self):
        dialog = EmailSablonDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.email_sablon_ekle(veri['ad'], veri['konu'], veri['icerik'], veri['kategori'])
            QMessageBox.information(self, "Basarili", "Sablon eklendi!")
            self.veriyi_yenile()
    
    def email_sablon_duzenle(self):
        row = self.sablon_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Sablon secin!")
            return
        
        sablon_id = int(self.sablon_tablo.item(row, 0).text())
        sablonlar = self.db.email_sablonlari_getir()
        sablon = next((s for s in sablonlar if s['id'] == sablon_id), None)
        
        dialog = EmailSablonDialog(self, sablon)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.email_sablon_sil(sablon_id)
            self.db.email_sablon_ekle(veri['ad'], veri['konu'], veri['icerik'], veri['kategori'])
            QMessageBox.information(self, "Basarili", "Sablon guncellendi!")
            self.veriyi_yenile()
    
    def email_sablon_sil(self):
        row = self.sablon_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Sablon secin!")
            return
        
        sablon_id = int(self.sablon_tablo.item(row, 0).text())
        ad = self.sablon_tablo.item(row, 1).text()
        
        cevap = QMessageBox.question(self, "Silme Onay", f"{ad} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.email_sablon_sil(sablon_id)
            QMessageBox.information(self, "Basarili", "Sablon silindi!")
            self.veriyi_yenile()
    
    def email_gonder(self):
        musteriler = self.db.musterileri_getir()
        sablonlar = self.db.email_sablonlari_getir()
        
        if not musteriler:
            QMessageBox.warning(self, "Uyari", "Musteri yok!")
            return
        
        dialog = EmailGonderDialog(self, musteriler, sablonlar)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            
            # Alici listesi belirle
            if veri['alici_tip'] == "Tek Musteri":
                hedef_musteriler = [m for m in musteriler if m['id'] == veri['musteri_id']]
            elif veri['alici_tip'] == "Tum Aktifler":
                hedef_musteriler = [m for m in musteriler if m['durum'] == 'Aktif']
            elif veri['alici_tip'] == "Tum Pasifler":
                hedef_musteriler = [m for m in musteriler if m['durum'] == 'Pasif']
            else:
                hedef_musteriler = musteriler
            
            # Her musteriye gonder
            gonderim_sayisi = 0
            for musteri in hedef_musteriler:
                konu = veri['konu'].replace("{ad}", musteri['ad']).replace("{email}", musteri['email']).replace("{sirket}", musteri['sirket'] or "")
                icerik = veri['icerik'].replace("{ad}", musteri['ad']).replace("{email}", musteri['email']).replace("{sirket}", musteri['sirket'] or "")
                
                self.db.email_gonder_kayit(musteri['id'], veri['sablon_id'], konu, icerik)
                gonderim_sayisi += 1
            
            QMessageBox.information(self, "Basarili", f"{gonderim_sayisi} email gonderildi!")
            self.veriyi_yenile()
    
    def email_yenile(self):
        if not hasattr(self, 'sablon_tablo'):
            return
        
        sablonlar = self.db.email_sablonlari_getir()
        gecmis = self.db.email_gecmis_getir()
        
        # KPI
        bugun = datetime.now().strftime('%Y-%m-%d')
        bugun_sayi = sum(1 for e in gecmis if e['tarih'].startswith(bugun))
        
        self.kpi_sablon_sayi.deger_label.setText(str(len(sablonlar)))
        self.kpi_gonderilen.deger_label.setText(str(len(gecmis)))
        self.kpi_bugun_gonderilen.deger_label.setText(str(bugun_sayi))
        
        # Sablon Tablo
        self.sablon_tablo.setRowCount(0)
        for s in sablonlar:
            row = self.sablon_tablo.rowCount()
            self.sablon_tablo.insertRow(row)
            
            self.sablon_tablo.setItem(row, 0, QTableWidgetItem(str(s['id'])))
            self.sablon_tablo.setItem(row, 1, QTableWidgetItem(s['ad']))
            
            kategori_item = QTableWidgetItem(s['kategori'])
            if s['kategori'] == 'Pazarlama':
                kategori_item.setForeground(QColor(COLORS['warning']))
            elif s['kategori'] == 'Hosgeldin':
                kategori_item.setForeground(QColor(COLORS['success']))
            elif s['kategori'] == 'Hatirlatma':
                kategori_item.setForeground(QColor(COLORS['danger']))
            else:
                kategori_item.setForeground(QColor(COLORS['secondary']))
            self.sablon_tablo.setItem(row, 2, kategori_item)
            
            self.sablon_tablo.setItem(row, 3, QTableWidgetItem(s['konu']))
            
            onizleme = s['icerik'][:60] + "..." if len(s['icerik']) > 60 else s['icerik']
            onizleme = onizleme.replace("\n", " ")
            self.sablon_tablo.setItem(row, 4, QTableWidgetItem(onizleme))
        
        # Gecmis Tablo
        self.email_gecmis_tablo.setRowCount(0)
        for e in gecmis[:100]:
            row = self.email_gecmis_tablo.rowCount()
            self.email_gecmis_tablo.insertRow(row)
            
            self.email_gecmis_tablo.setItem(row, 0, QTableWidgetItem(str(e['id'])))
            self.email_gecmis_tablo.setItem(row, 1, QTableWidgetItem(e['musteri_adi']))
            self.email_gecmis_tablo.setItem(row, 2, QTableWidgetItem(e['musteri_email']))
            self.email_gecmis_tablo.setItem(row, 3, QTableWidgetItem(e['konu']))
            self.email_gecmis_tablo.setItem(row, 4, QTableWidgetItem(e['tarih']))
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 5: GELISMIS ANALITIK (Trend + Lineer Regresyon)
    # ─────────────────────────────────────────────────────────────────────────
    
    def analitik_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("Gelismis Analitik")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        # Insight Kartlari (3 adet)
        insight_layout = QHBoxLayout()
        insight_layout.setSpacing(15)
        
        self.insight_trend = self.kpi_kart("Buyume Trendi", "0%", COLORS['success'])
        self.insight_ortalama = self.kpi_kart("Aylik Ortalama", "0 TRY", COLORS['primary'])
        self.insight_tahmin = self.kpi_kart("Sonraki Ay Tahmini", "0 TRY", COLORS['warning'])
        
        insight_layout.addWidget(self.insight_trend)
        insight_layout.addWidget(self.insight_ortalama)
        insight_layout.addWidget(self.insight_tahmin)
        
        layout.addLayout(insight_layout)
        
        # 2 Grafik Yan Yana
        grafik_layout = QHBoxLayout()
        grafik_layout.setSpacing(15)
        
        # Sol: Gelir Trendi + Regresyon
        self.regresyon_grafik = self.grafik_widget()
        regresyon_frame = QFrame()
        regresyon_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        regresyon_vlayout = QVBoxLayout()
        regresyon_baslik = QLabel("Gelir Trendi + Lineer Regresyon")
        regresyon_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        regresyon_vlayout.addWidget(regresyon_baslik)
        regresyon_vlayout.addWidget(self.regresyon_grafik)
        regresyon_frame.setLayout(regresyon_vlayout)
        
        # Sag: Urun Bazli Satis
        self.urun_grafik = self.grafik_widget()
        urun_frame = QFrame()
        urun_frame.setStyleSheet(f"border: 1px solid {COLORS['border']}; border-radius: 8px;")
        urun_vlayout = QVBoxLayout()
        urun_baslik = QLabel("En Cok Satan Urunler")
        urun_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        urun_vlayout.addWidget(urun_baslik)
        urun_vlayout.addWidget(self.urun_grafik)
        urun_frame.setLayout(urun_vlayout)
        
        grafik_layout.addWidget(regresyon_frame)
        grafik_layout.addWidget(urun_frame)
        
        layout.addLayout(grafik_layout)
        
        # 2 Tablo Yan Yana
        tablo_layout = QHBoxLayout()
        tablo_layout.setSpacing(15)
        
        # Sol Tablo: Top Musteriler
        sol_widget = QWidget()
        sol_layout = QVBoxLayout()
        sol_layout.setContentsMargins(0, 0, 0, 0)
        sol_baslik = QLabel("Top 10 Musteri (Harcama)")
        sol_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        
        self.top_musteri_tablo = QTableWidget()
        self.top_musteri_tablo.setColumnCount(4)
        self.top_musteri_tablo.setHorizontalHeaderLabels(["Sira", "Musteri", "Satis Sayisi", "Toplam"])
        self.top_musteri_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.top_musteri_tablo.setMaximumHeight(300)
        
        sol_layout.addWidget(sol_baslik)
        sol_layout.addWidget(self.top_musteri_tablo)
        sol_widget.setLayout(sol_layout)
        
        # Sag Tablo: Aylik Trend
        sag_widget = QWidget()
        sag_layout = QVBoxLayout()
        sag_layout.setContentsMargins(0, 0, 0, 0)
        sag_baslik = QLabel("Aylik Detayli Analiz")
        sag_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        
        self.aylik_analiz_tablo = QTableWidget()
        self.aylik_analiz_tablo.setColumnCount(4)
        self.aylik_analiz_tablo.setHorizontalHeaderLabels(["Ay", "Gelir", "Satis", "Buyume %"])
        self.aylik_analiz_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.aylik_analiz_tablo.setMaximumHeight(300)
        
        sag_layout.addWidget(sag_baslik)
        sag_layout.addWidget(self.aylik_analiz_tablo)
        sag_widget.setLayout(sag_layout)
        
        tablo_layout.addWidget(sol_widget)
        tablo_layout.addWidget(sag_widget)
        
        layout.addLayout(tablo_layout)
        
        # Yenile butonu
        yenile_btn = QPushButton("Analitik Verileri Yenile")
        yenile_btn.clicked.connect(self.analitik_yenile)
        layout.addWidget(yenile_btn)
        
        widget.setLayout(layout)
        return widget
    
    def lineer_regresyon(self, x_list, y_list):
        """Basit lineer regresyon: y = ax + b"""
        n = len(x_list)
        if n < 2:
            return 0, 0
        
        x_ort = sum(x_list) / n
        y_ort = sum(y_list) / n
        
        pay = sum((x_list[i] - x_ort) * (y_list[i] - y_ort) for i in range(n))
        payda = sum((x_list[i] - x_ort) ** 2 for i in range(n))
        
        if payda == 0:
            return 0, y_ort
        
        a = pay / payda  # egim
        b = y_ort - a * x_ort  # kesisim
        
        return a, b
    
    def analitik_yenile(self):
        if not hasattr(self, 'regresyon_grafik'):
            return
        
        satislar = self.db.satislari_getir()
        
        # AYLIK GELIR HESAPLA
        aylar = {}
        for satis in satislar:
            ay = satis['tarih'][:7]
            if ay not in aylar:
                aylar[ay] = {'gelir': 0, 'satis': 0}
            aylar[ay]['gelir'] += satis['toplam']
            aylar[ay]['satis'] += 1
        
        sorted_aylar = sorted(aylar.keys())
        gelirler = [aylar[a]['gelir'] for a in sorted_aylar]
        satis_sayilari = [aylar[a]['satis'] for a in sorted_aylar]
        
        # ─── INSIGHT KARTLARI ───
        if len(gelirler) >= 2:
            son_ay = gelirler[-1]
            onceki_ay = gelirler[-2]
            buyume = ((son_ay - onceki_ay) / onceki_ay * 100) if onceki_ay > 0 else 0
            self.insight_trend.deger_label.setText(f"{buyume:+.1f}%")
            
            if buyume >= 0:
                self.insight_trend.deger_label.setStyleSheet(f"color: {COLORS['success']};")
            else:
                self.insight_trend.deger_label.setStyleSheet(f"color: {COLORS['danger']};")
        else:
            self.insight_trend.deger_label.setText("N/A")
        
        if gelirler:
            ortalama = sum(gelirler) / len(gelirler)
            self.insight_ortalama.deger_label.setText(f"{ortalama:,.0f} TRY")
        
        # LINEER REGRESYON ile TAHMIN
        if len(gelirler) >= 2:
            x_list = list(range(len(gelirler)))
            a, b = self.lineer_regresyon(x_list, gelirler)
            tahmin = a * len(gelirler) + b
            self.insight_tahmin.deger_label.setText(f"{max(0, tahmin):,.0f} TRY")
        else:
            self.insight_tahmin.deger_label.setText("N/A")
        
        # ─── REGRESYON GRAFIGI ───
        self.regresyon_grafik.fig.clear()
        ax = self.regresyon_grafik.fig.add_subplot(111)
        
        if len(gelirler) >= 2:
            x_list = list(range(len(gelirler)))
            a, b = self.lineer_regresyon(x_list, gelirler)
            
            # Gercek veri (bar)
            ax.bar(x_list, gelirler, color=COLORS['primary'], alpha=0.6, label='Gercek')
            
            # Regresyon dogrusu
            regresyon_y = [a * x + b for x in x_list]
            ax.plot(x_list, regresyon_y, color=COLORS['warning'], linewidth=2, label='Trend', marker='o')
            
            # Tahmin noktasi
            tahmin_x = len(gelirler)
            tahmin_y = max(0, a * tahmin_x + b)
            ax.scatter([tahmin_x], [tahmin_y], color=COLORS['success'], s=150, zorder=5, label='Tahmin')
            
            ax.set_xticks(list(x_list) + [tahmin_x])
            etiketler = [a[5:] for a in sorted_aylar] + ['Sonraki']
            ax.set_xticklabels(etiketler, color=COLORS['text_sec'], rotation=45)
            ax.set_ylabel('Gelir (TRY)', color=COLORS['text_sec'])
            ax.tick_params(colors=COLORS['text_sec'])
            ax.legend(facecolor=COLORS['bg_secondary'], edgecolor=COLORS['border'], labelcolor=COLORS['text_main'])
        
        ax.set_facecolor(COLORS['bg_secondary'])
        ax.spines['bottom'].set_color(COLORS['border'])
        ax.spines['left'].set_color(COLORS['border'])
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.1)
        
        self.regresyon_grafik.fig.tight_layout()
        self.regresyon_grafik.draw()
        
        # ─── URUN GRAFIGI (Horizontal Bar) ───
        self.urun_grafik.fig.clear()
        ax2 = self.urun_grafik.fig.add_subplot(111)
        
        urun_say = {}
        for satis in satislar:
            urun = satis['urun']
            urun_say[urun] = urun_say.get(urun, 0) + satis['toplam']
        
        if urun_say:
            sorted_urunler = sorted(urun_say.items(), key=lambda x: x[1], reverse=True)[:6]
            urunler = [u[0] for u in sorted_urunler]
            gelir_urun = [u[1] for u in sorted_urunler]
            
            ax2.barh(range(len(urunler)), gelir_urun, color=COLORS['primary'], alpha=0.8)
            ax2.set_yticks(range(len(urunler)))
            ax2.set_yticklabels(urunler, color=COLORS['text_sec'], fontsize=9)
            ax2.set_xlabel('Toplam Gelir (TRY)', color=COLORS['text_sec'])
            ax2.tick_params(colors=COLORS['text_sec'])
            ax2.invert_yaxis()
        
        ax2.set_facecolor(COLORS['bg_secondary'])
        ax2.spines['bottom'].set_color(COLORS['border'])
        ax2.spines['left'].set_color(COLORS['border'])
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
        ax2.grid(True, alpha=0.1, axis='x')
        
        self.urun_grafik.fig.tight_layout()
        self.urun_grafik.draw()
        
        # ─── TOP MUSTERI TABLOSU ───
        musteri_harcama = {}
        for satis in satislar:
            ad = satis['musteri_adi']
            if ad not in musteri_harcama:
                musteri_harcama[ad] = {'toplam': 0, 'sayi': 0}
            musteri_harcama[ad]['toplam'] += satis['toplam']
            musteri_harcama[ad]['sayi'] += 1
        
        sorted_musteriler = sorted(musteri_harcama.items(), key=lambda x: x[1]['toplam'], reverse=True)[:10]
        
        self.top_musteri_tablo.setRowCount(0)
        for i, (ad, data) in enumerate(sorted_musteriler, 1):
            row = self.top_musteri_tablo.rowCount()
            self.top_musteri_tablo.insertRow(row)
            
            sira_item = QTableWidgetItem(f"#{i}")
            if i == 1:
                sira_item.setForeground(QColor(COLORS['success']))
            elif i <= 3:
                sira_item.setForeground(QColor(COLORS['warning']))
            self.top_musteri_tablo.setItem(row, 0, sira_item)
            
            self.top_musteri_tablo.setItem(row, 1, QTableWidgetItem(ad))
            self.top_musteri_tablo.setItem(row, 2, QTableWidgetItem(str(data['sayi'])))
            self.top_musteri_tablo.setItem(row, 3, QTableWidgetItem(f"{data['toplam']:,.0f} TRY"))
        
        # ─── AYLIK ANALIZ TABLOSU ───
        self.aylik_analiz_tablo.setRowCount(0)
        for i, ay in enumerate(sorted_aylar):
            row = self.aylik_analiz_tablo.rowCount()
            self.aylik_analiz_tablo.insertRow(row)
            
            self.aylik_analiz_tablo.setItem(row, 0, QTableWidgetItem(ay))
            self.aylik_analiz_tablo.setItem(row, 1, QTableWidgetItem(f"{aylar[ay]['gelir']:,.0f} TRY"))
            self.aylik_analiz_tablo.setItem(row, 2, QTableWidgetItem(str(aylar[ay]['satis'])))
            
            if i > 0:
                onceki = aylar[sorted_aylar[i-1]]['gelir']
                if onceki > 0:
                    buyume = (aylar[ay]['gelir'] - onceki) / onceki * 100
                    buyume_item = QTableWidgetItem(f"{buyume:+.1f}%")
                    if buyume >= 0:
                        buyume_item.setForeground(QColor(COLORS['success']))
                    else:
                        buyume_item.setForeground(QColor(COLORS['danger']))
                    self.aylik_analiz_tablo.setItem(row, 3, buyume_item)
                else:
                    self.aylik_analiz_tablo.setItem(row, 3, QTableWidgetItem("N/A"))
            else:
                self.aylik_analiz_tablo.setItem(row, 3, QTableWidgetItem("-"))
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 4: ODULLER & PUAN SISTEMI
    # ─────────────────────────────────────────────────────────────────────────
    
    def oduller_tab_olustur(self):
        widget = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        
        # 3 KPI Karti
        kpi_layout = QHBoxLayout()
        kpi_layout.setSpacing(15)
        
        self.kpi_toplam_puan = self.kpi_kart("Toplam Puan", "0", COLORS['primary'])
        self.kpi_platin = self.kpi_kart("Platin Musteri", "0", COLORS['success'])
        self.kpi_hediye = self.kpi_kart("Verilen Hediye", "0", COLORS['warning'])
        
        kpi_layout.addWidget(self.kpi_toplam_puan)
        kpi_layout.addWidget(self.kpi_platin)
        kpi_layout.addWidget(self.kpi_hediye)
        
        main_layout.addLayout(kpi_layout)
        
        # 2 alt panel: Puan Listesi + Odul Katalog
        alt_layout = QHBoxLayout()
        alt_layout.setSpacing(15)
        
        # SOL: PUAN LISTESI
        sol_widget = QWidget()
        sol_layout = QVBoxLayout()
        sol_layout.setContentsMargins(0, 0, 0, 0)
        
        sol_baslik_layout = QHBoxLayout()
        sol_baslik = QLabel("Musteri Puanlari")
        sol_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        puan_ekle_btn = QPushButton("Puan Islemi")
        puan_ekle_btn.clicked.connect(self.puan_islemi_yap)
        hediye_btn = QPushButton("Hediye Ver")
        hediye_btn.clicked.connect(self.hediye_ver_yap)
        
        sol_baslik_layout.addWidget(sol_baslik)
        sol_baslik_layout.addStretch()
        sol_baslik_layout.addWidget(puan_ekle_btn)
        sol_baslik_layout.addWidget(hediye_btn)
        
        self.puan_tablo = QTableWidget()
        self.puan_tablo.setColumnCount(4)
        self.puan_tablo.setHorizontalHeaderLabels(["Musteri", "Puan", "Seviye", "Son Guncelleme"])
        self.puan_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        sol_layout.addLayout(sol_baslik_layout)
        sol_layout.addWidget(self.puan_tablo)
        sol_widget.setLayout(sol_layout)
        
        # SAG: ODUL KATALOG
        sag_widget = QWidget()
        sag_layout = QVBoxLayout()
        sag_layout.setContentsMargins(0, 0, 0, 0)
        
        sag_baslik_layout = QHBoxLayout()
        sag_baslik = QLabel("Odul Katalogu")
        sag_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        odul_ekle_btn = QPushButton("Odul Ekle")
        odul_ekle_btn.clicked.connect(self.odul_ekle)
        odul_sil_btn = QPushButton("Odul Sil")
        odul_sil_btn.clicked.connect(self.odul_sil)
        
        sag_baslik_layout.addWidget(sag_baslik)
        sag_baslik_layout.addStretch()
        sag_baslik_layout.addWidget(odul_ekle_btn)
        sag_baslik_layout.addWidget(odul_sil_btn)
        
        self.odul_tablo = QTableWidget()
        self.odul_tablo.setColumnCount(5)
        self.odul_tablo.setHorizontalHeaderLabels(["ID", "Ad", "Aciklama", "Puan", "Stok"])
        self.odul_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        sag_layout.addLayout(sag_baslik_layout)
        sag_layout.addWidget(self.odul_tablo)
        sag_widget.setLayout(sag_layout)
        
        alt_layout.addWidget(sol_widget)
        alt_layout.addWidget(sag_widget)
        
        main_layout.addLayout(alt_layout)
        
        # En altta: Verilen Hediyeler
        main_layout.addWidget(QLabel("Verilen Hediyeler"))
        
        self.hediye_tablo = QTableWidget()
        self.hediye_tablo.setColumnCount(5)
        self.hediye_tablo.setHorizontalHeaderLabels(["ID", "Musteri", "Hediye", "Puan", "Tarih"])
        self.hediye_tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.hediye_tablo.setMaximumHeight(200)
        
        main_layout.addWidget(self.hediye_tablo)
        
        widget.setLayout(main_layout)
        return widget
    
    def odul_ekle(self):
        dialog = OdulKatalogDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.odul_ekle_katalog(veri['ad'], veri['aciklama'], veri['puan'], veri['stok'])
            QMessageBox.information(self, "Basarili", "Odul eklendi!")
            self.veriyi_yenile()
    
    def odul_sil(self):
        row = self.odul_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Odul secin!")
            return
        
        odul_id = int(self.odul_tablo.item(row, 0).text())
        ad = self.odul_tablo.item(row, 1).text()
        
        cevap = QMessageBox.question(self, "Silme Onay", f"{ad} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.odul_sil(odul_id)
            QMessageBox.information(self, "Basarili", "Odul silindi!")
            self.veriyi_yenile()
    
    def puan_islemi_yap(self):
        row = self.puan_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Musteri secin!")
            return
        
        musteri_adi = self.puan_tablo.item(row, 0).text()
        puanlar = self.db.puanlari_getir()
        musteri_puan = puanlar[row]
        musteri_id = musteri_puan['musteri_id']
        
        dialog = PuanEkleDialog(self, musteri_adi)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            if veri['islem'] == "Puan Ekle":
                self.db.puan_ekle(musteri_id, veri['miktar'])
                QMessageBox.information(self, "Basarili", f"{veri['miktar']} puan eklendi!")
            else:
                if self.db.puan_dusur(musteri_id, veri['miktar']):
                    QMessageBox.information(self, "Basarili", f"{veri['miktar']} puan cikarildi!")
                else:
                    QMessageBox.warning(self, "Hata", "Yeterli puan yok!")
            self.veriyi_yenile()
    
    def hediye_ver_yap(self):
        musteriler = self.db.musterileri_getir()
        oduller = self.db.odul_katalog_getir()
        
        if not musteriler:
            QMessageBox.warning(self, "Uyari", "Musteri yok!")
            return
        if not oduller:
            QMessageBox.warning(self, "Uyari", "Odul katalogu bos!")
            return
        
        dialog = HediyeVerDialog(self, musteriler, oduller)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            basari, mesaj = self.db.hediye_ver(veri['musteri_id'], veri['odul_id'])
            if basari:
                QMessageBox.information(self, "Basarili", "Hediye verildi!")
            else:
                QMessageBox.warning(self, "Hata", mesaj)
            self.veriyi_yenile()
    
    # ─────────────────────────────────────────────────────────────────────────
    # TIER 3: RAPORLAR
    # ─────────────────────────────────────────────────────────────────────────
    
    def raporlar_tab_olustur(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("Raporlar ve Export")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        layout.addWidget(title)
        
        layout.addWidget(QLabel("Musteri Raporu"))
        musteri_layout = QHBoxLayout()
        musteri_csv_btn = QPushButton("CSV Indir")
        musteri_csv_btn.clicked.connect(lambda: self.rapor_indir("musteri", "csv"))
        musteri_excel_btn = QPushButton("Excel Indir")
        musteri_excel_btn.clicked.connect(lambda: self.rapor_indir("musteri", "excel"))
        musteri_layout.addWidget(musteri_csv_btn)
        musteri_layout.addWidget(musteri_excel_btn)
        musteri_layout.addStretch()
        layout.addLayout(musteri_layout)
        
        layout.addWidget(QLabel("Satis Raporu"))
        satis_layout = QHBoxLayout()
        satis_csv_btn = QPushButton("CSV Indir")
        satis_csv_btn.clicked.connect(lambda: self.rapor_indir("satis", "csv"))
        satis_excel_btn = QPushButton("Excel Indir")
        satis_excel_btn.clicked.connect(lambda: self.rapor_indir("satis", "excel"))
        satis_layout.addWidget(satis_csv_btn)
        satis_layout.addWidget(satis_excel_btn)
        satis_layout.addStretch()
        layout.addLayout(satis_layout)
        
        layout.addWidget(QLabel("Destek Raporu"))
        destek_layout = QHBoxLayout()
        destek_csv_btn = QPushButton("CSV Indir")
        destek_csv_btn.clicked.connect(lambda: self.rapor_indir("destek", "csv"))
        destek_excel_btn = QPushButton("Excel Indir")
        destek_excel_btn.clicked.connect(lambda: self.rapor_indir("destek", "excel"))
        destek_layout.addWidget(destek_csv_btn)
        destek_layout.addWidget(destek_excel_btn)
        destek_layout.addStretch()
        layout.addLayout(destek_layout)
        
        layout.addStretch()
        
        widget.setLayout(layout)
        return widget
    
    def rapor_indir(self, rapor_tipi, format_tipi):
        dialog = RaporDialog(self, rapor_tipi)
        if dialog.exec_() != QDialog.Accepted:
            return
        
        veri = dialog.veri_al()
        basla = veri['basla']
        bitis = veri['bitis']
        
        if rapor_tipi == "musteri":
            self.musteri_raporu_indir(basla, bitis, format_tipi)
        elif rapor_tipi == "satis":
            self.satis_raporu_indir(basla, bitis, format_tipi)
        elif rapor_tipi == "destek":
            self.destek_raporu_indir(basla, bitis, format_tipi)
    
    def musteri_raporu_indir(self, basla, bitis, format_tipi):
        musteriler = self.db.musterileri_getir()
        
        filtrelenmis = [m for m in musteriler if basla <= m['katilim_tarihi'] <= bitis]
        
        if format_tipi == "csv":
            dosya = QFileDialog.getSaveFileName(self, "CSV Kaydet", "musteri_raporu.csv", "CSV Files (*.csv)")[0]
            if not dosya:
                return
            
            with open(dosya, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Ad', 'Email', 'Telefon', 'Sirket', 'Durum', 'Toplam Harcama', 'Katilim Tarihi'])
                for m in filtrelenmis:
                    writer.writerow([m['id'], m['ad'], m['email'], m['telefon'], m['sirket'], m['durum'], m['toplam_harcama'], m['katilim_tarihi']])
            
            QMessageBox.information(self, "Basarili", f"Rapor kaydedildi: {dosya}")
        
        elif format_tipi == "excel" and OPENPYXL_AVAILABLE:
            dosya = QFileDialog.getSaveFileName(self, "Excel Kaydet", "musteri_raporu.xlsx", "Excel Files (*.xlsx)")[0]
            if not dosya:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Musteriler"
            
            headers = ['ID', 'Ad', 'Email', 'Telefon', 'Sirket', 'Durum', 'Toplam Harcama', 'Katilim Tarihi']
            ws.append(headers)
            
            header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
            header_font = Font(color="ffffff", bold=True)
            
            for col in ws[1]:
                col.fill = header_fill
                col.font = header_font
                col.alignment = Alignment(horizontal='center')
            
            for m in filtrelenmis:
                ws.append([m['id'], m['ad'], m['email'], m['telefon'], m['sirket'], m['durum'], m['toplam_harcama'], m['katilim_tarihi']])
            
            wb.save(dosya)
            QMessageBox.information(self, "Basarili", f"Rapor kaydedildi: {dosya}")
        
        else:
            QMessageBox.warning(self, "Uyari", "openpyxl kutuphane kurulu degil!")
    
    def satis_raporu_indir(self, basla, bitis, format_tipi):
        satislar = self.db.satislari_getir()
        
        filtrelenmis = [s for s in satislar if basla <= s['tarih'] <= bitis]
        
        if format_tipi == "csv":
            dosya = QFileDialog.getSaveFileName(self, "CSV Kaydet", "satis_raporu.csv", "CSV Files (*.csv)")[0]
            if not dosya:
                return
            
            with open(dosya, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Musteri', 'Urun', 'Tutar', 'Adet', 'Toplam', 'Durum', 'Tarih'])
                for s in filtrelenmis:
                    writer.writerow([s['id'], s['musteri_adi'], s['urun'], s['tutar'], s['adet'], s['toplam'], s['durum'], s['tarih']])
            
            QMessageBox.information(self, "Basarili", f"Rapor kaydedildi: {dosya}")
        
        elif format_tipi == "excel" and OPENPYXL_AVAILABLE:
            dosya = QFileDialog.getSaveFileName(self, "Excel Kaydet", "satis_raporu.xlsx", "Excel Files (*.xlsx)")[0]
            if not dosya:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Satislar"
            
            headers = ['ID', 'Musteri', 'Urun', 'Tutar', 'Adet', 'Toplam', 'Durum', 'Tarih']
            ws.append(headers)
            
            header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
            header_font = Font(color="ffffff", bold=True)
            
            for col in ws[1]:
                col.fill = header_fill
                col.font = header_font
                col.alignment = Alignment(horizontal='center')
            
            for s in filtrelenmis:
                ws.append([s['id'], s['musteri_adi'], s['urun'], s['tutar'], s['adet'], s['toplam'], s['durum'], s['tarih']])
            
            wb.save(dosya)
            QMessageBox.information(self, "Basarili", f"Rapor kaydedildi: {dosya}")
        
        else:
            QMessageBox.warning(self, "Uyari", "openpyxl kutuphane kurulu degil!")
    
    def destek_raporu_indir(self, basla, bitis, format_tipi):
        destekler = self.db.destekleri_getir()
        
        filtrelenmis = [d for d in destekler if basla <= d['tarih'] <= bitis]
        
        if format_tipi == "csv":
            dosya = QFileDialog.getSaveFileName(self, "CSV Kaydet", "destek_raporu.csv", "CSV Files (*.csv)")[0]
            if not dosya:
                return
            
            with open(dosya, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Musteri', 'Konu', 'Aciklama', 'Oncelik', 'Durum', 'Tarih'])
                for d in filtrelenmis:
                    writer.writerow([d['id'], d['musteri_adi'], d['konu'], d['aciklama'], d['oncelik'], d['durum'], d['tarih']])
            
            QMessageBox.information(self, "Basarili", f"Rapor kaydedildi: {dosya}")
        
        elif format_tipi == "excel" and OPENPYXL_AVAILABLE:
            dosya = QFileDialog.getSaveFileName(self, "Excel Kaydet", "destek_raporu.xlsx", "Excel Files (*.xlsx)")[0]
            if not dosya:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Destek"
            
            headers = ['ID', 'Musteri', 'Konu', 'Aciklama', 'Oncelik', 'Durum', 'Tarih']
            ws.append(headers)
            
            header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
            header_font = Font(color="ffffff", bold=True)
            
            for col in ws[1]:
                col.fill = header_fill
                col.font = header_font
                col.alignment = Alignment(horizontal='center')
            
            for d in filtrelenmis:
                ws.append([d['id'], d['musteri_adi'], d['konu'], d['aciklama'], d['oncelik'], d['durum'], d['tarih']])
            
            wb.save(dosya)
            QMessageBox.information(self, "Basarili", f"Rapor kaydedildi: {dosya}")
        
        else:
            QMessageBox.warning(self, "Uyari", "openpyxl kutuphane kurulu degil!")
    
    # ─────────────────────────────────────────────────────────────────────────
    # CRUD İŞLEMLERİ
    # ─────────────────────────────────────────────────────────────────────────
    
    def musteri_ekle(self):
        dialog = MusteriDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.musteri_ekle(veri['ad'], veri['email'], veri['telefon'], veri['sirket'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Musteri eklendi!")
            self.veriyi_yenile()
    
    def musteri_duzenle(self):
        row = self.musteri_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Musteri secin!")
            return
        
        musteri_id = int(self.musteri_tablo.item(row, 0).text())
        musteriler = self.db.musterileri_getir()
        musteri = next((m for m in musteriler if m['id'] == musteri_id), None)
        
        dialog = MusteriDialog(self, musteri)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.musteri_guncelle(musteri_id, veri['ad'], veri['email'], veri['telefon'], veri['sirket'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Musteri guncellendi!")
            self.veriyi_yenile()
    
    def musteri_sil(self):
        row = self.musteri_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Musteri secin!")
            return
        
        musteri_id = int(self.musteri_tablo.item(row, 0).text())
        ad = self.musteri_tablo.item(row, 1).text()
        
        cevap = QMessageBox.question(self, "Silme Onay", f"{ad} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.musteri_sil(musteri_id)
            QMessageBox.information(self, "Basarili", "Musteri silindi!")
            self.veriyi_yenile()
    
    def musterileri_ara(self):
        arama = self.musteri_arama.text().lower()
        for row in range(self.musteri_tablo.rowCount()):
            visible = False
            for col in range(1, 5):
                item = self.musteri_tablo.item(row, col)
                if item and arama in item.text().lower():
                    visible = True
                    break
            self.musteri_tablo.setRowHidden(row, not visible)
    
    def satis_ekle(self):
        musteriler = self.db.musterileri_getir()
        if not musteriler:
            QMessageBox.warning(self, "Uyari", "Once musteri ekleyin!")
            return
        
        dialog = SatisDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            musteri_id = musteriler[0]['id']
            self.db.satis_ekle(musteri_id, veri['urun'], veri['tutar'], veri['adet'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Satis eklendi!")
            self.veriyi_yenile()
    
    def satis_duzenle(self):
        row = self.satis_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Satis secin!")
            return
        
        satis_id = int(self.satis_tablo.item(row, 0).text())
        satislar = self.db.satislari_getir()
        satis = next((s for s in satislar if s['id'] == satis_id), None)
        
        dialog = SatisDialog(self, satis)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.satis_guncelle(satis_id, veri['urun'], veri['tutar'], veri['adet'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Satis guncellendi!")
            self.veriyi_yenile()
    
    def satis_sil(self):
        row = self.satis_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Satis secin!")
            return
        
        satis_id = int(self.satis_tablo.item(row, 0).text())
        cevap = QMessageBox.question(self, "Silme Onay", f"Satis #{satis_id} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.satis_sil(satis_id)
            QMessageBox.information(self, "Basarili", "Satis silindi!")
            self.veriyi_yenile()
    
    def satislari_ara(self):
        arama = self.satis_arama.text().lower()
        for row in range(self.satis_tablo.rowCount()):
            visible = False
            for col in range(1, 7):
                item = self.satis_tablo.item(row, col)
                if item and arama in item.text().lower():
                    visible = True
                    break
            self.satis_tablo.setRowHidden(row, not visible)
    
    def destek_ekle(self):
        musteriler = self.db.musterileri_getir()
        if not musteriler:
            QMessageBox.warning(self, "Uyari", "Once musteri ekleyin!")
            return
        
        dialog = DestekDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            musteri_id = musteriler[0]['id']
            self.db.destek_ekle(musteri_id, veri['konu'], veri['aciklama'], veri['oncelik'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Destek eklendi!")
            self.veriyi_yenile()
    
    def destek_duzenle(self):
        row = self.destek_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Destek secin!")
            return
        
        destek_id = int(self.destek_tablo.item(row, 0).text())
        destekler = self.db.destekleri_getir()
        destek = next((d for d in destekler if d['id'] == destek_id), None)
        
        dialog = DestekDialog(self, destek)
        if dialog.exec_() == QDialog.Accepted:
            veri = dialog.veri_al()
            self.db.destek_guncelle(destek_id, veri['oncelik'], veri['durum'])
            QMessageBox.information(self, "Basarili", "Destek guncellendi!")
            self.veriyi_yenile()
    
    def destek_sil(self):
        row = self.destek_tablo.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyari", "Destek secin!")
            return
        
        destek_id = int(self.destek_tablo.item(row, 0).text())
        cevap = QMessageBox.question(self, "Silme Onay", f"Destek #{destek_id} silinsin mi?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            self.db.destek_sil(destek_id)
            QMessageBox.information(self, "Basarili", "Destek silindi!")
            self.veriyi_yenile()
    
    def destekleri_ara(self):
        arama = self.destek_arama.text().lower()
        for row in range(self.destek_tablo.rowCount()):
            visible = False
            for col in range(1, 7):
                item = self.destek_tablo.item(row, col)
                if item and arama in item.text().lower():
                    visible = True
                    break
            self.destek_tablo.setRowHidden(row, not visible)
    
    def veriyi_yenile(self):
        musteriler = self.db.musterileri_getir()
        satislar = self.db.satislari_getir()
        destekler = self.db.destekleri_getir()
        
        self.musteri_tablo.setRowCount(0)
        for musteri in musteriler:
            row = self.musteri_tablo.rowCount()
            self.musteri_tablo.insertRow(row)
            
            self.musteri_tablo.setItem(row, 0, QTableWidgetItem(str(musteri['id'])))
            self.musteri_tablo.setItem(row, 1, QTableWidgetItem(musteri['ad']))
            self.musteri_tablo.setItem(row, 2, QTableWidgetItem(musteri['email']))
            self.musteri_tablo.setItem(row, 3, QTableWidgetItem(musteri['telefon'] or ""))
            self.musteri_tablo.setItem(row, 4, QTableWidgetItem(musteri['sirket'] or ""))
            
            durum_item = QTableWidgetItem(musteri['durum'])
            if musteri['durum'] == 'Aktif':
                durum_item.setForeground(QColor(COLORS['success']))
            elif musteri['durum'] == 'Pasif':
                durum_item.setForeground(QColor(COLORS['danger']))
            else:
                durum_item.setForeground(QColor(COLORS['warning']))
            self.musteri_tablo.setItem(row, 5, durum_item)
            
            self.musteri_tablo.setItem(row, 6, QTableWidgetItem(f"{musteri['toplam_harcama']:,.0f} TRY"))
        
        self.satis_tablo.setRowCount(0)
        for satis in satislar:
            row = self.satis_tablo.rowCount()
            self.satis_tablo.insertRow(row)
            
            self.satis_tablo.setItem(row, 0, QTableWidgetItem(str(satis['id'])))
            self.satis_tablo.setItem(row, 1, QTableWidgetItem(satis['musteri_adi']))
            self.satis_tablo.setItem(row, 2, QTableWidgetItem(satis['urun']))
            self.satis_tablo.setItem(row, 3, QTableWidgetItem(f"{satis['tutar']:,.0f} TRY"))
            self.satis_tablo.setItem(row, 4, QTableWidgetItem(str(satis['adet'])))
            self.satis_tablo.setItem(row, 5, QTableWidgetItem(f"{satis['toplam']:,.0f} TRY"))
            
            durum_item = QTableWidgetItem(satis['durum'])
            if satis['durum'] == 'Tamamlandi':
                durum_item.setForeground(QColor(COLORS['success']))
            elif satis['durum'] == 'Isleniysor':
                durum_item.setForeground(QColor(COLORS['warning']))
            else:
                durum_item.setForeground(QColor(COLORS['danger']))
            self.satis_tablo.setItem(row, 6, durum_item)
        
        self.destek_tablo.setRowCount(0)
        for destek in destekler:
            row = self.destek_tablo.rowCount()
            self.destek_tablo.insertRow(row)
            
            self.destek_tablo.setItem(row, 0, QTableWidgetItem(str(destek['id'])))
            self.destek_tablo.setItem(row, 1, QTableWidgetItem(destek['musteri_adi']))
            self.destek_tablo.setItem(row, 2, QTableWidgetItem(destek['konu']))
            self.destek_tablo.setItem(row, 3, QTableWidgetItem(destek['aciklama'][:50] if destek['aciklama'] else ""))
            
            oncelik_item = QTableWidgetItem(destek['oncelik'])
            if destek['oncelik'] == 'Acil':
                oncelik_item.setForeground(QColor(COLORS['danger']))
            elif destek['oncelik'] == 'Yuksek':
                oncelik_item.setForeground(QColor(COLORS['warning']))
            self.destek_tablo.setItem(row, 4, oncelik_item)
            
            durum_item = QTableWidgetItem(destek['durum'])
            if destek['durum'] == 'Cozuldu':
                durum_item.setForeground(QColor(COLORS['success']))
            self.destek_tablo.setItem(row, 5, durum_item)
            
            self.destek_tablo.setItem(row, 6, QTableWidgetItem(destek['tarih']))
        
        self.dashboard_yenile()
        self.oduller_yenile()
        self.analitik_yenile()
        self.email_yenile()
        self.kampanya_yenile()
        self.firsat_yenile()
        self.etkinlik_yenile()
        self.profil_combo_doldur()
        self.profil_yukle()
    
    def oduller_yenile(self):
        if not hasattr(self, 'puan_tablo'):
            return
        
        puanlar = self.db.puanlari_getir()
        oduller = self.db.odul_katalog_getir()
        hediyeler = self.db.hediyeleri_getir()
        
        # KPI
        toplam_puan = sum(p['puan'] for p in puanlar)
        platin_sayi = sum(1 for p in puanlar if p['seviye'] == 'Platin')
        hediye_sayi = len(hediyeler)
        
        self.kpi_toplam_puan.deger_label.setText(f"{toplam_puan:,}")
        self.kpi_platin.deger_label.setText(str(platin_sayi))
        self.kpi_hediye.deger_label.setText(str(hediye_sayi))
        
        # Puan Tablosu
        self.puan_tablo.setRowCount(0)
        for p in puanlar:
            row = self.puan_tablo.rowCount()
            self.puan_tablo.insertRow(row)
            
            self.puan_tablo.setItem(row, 0, QTableWidgetItem(p['musteri_adi']))
            self.puan_tablo.setItem(row, 1, QTableWidgetItem(f"{p['puan']:,}"))
            
            seviye_item = QTableWidgetItem(p['seviye'])
            if p['seviye'] == 'Platin':
                seviye_item.setForeground(QColor(COLORS['success']))
            elif p['seviye'] == 'Altin':
                seviye_item.setForeground(QColor(COLORS['warning']))
            elif p['seviye'] == 'Gumus':
                seviye_item.setForeground(QColor(COLORS['secondary']))
            else:
                seviye_item.setForeground(QColor(COLORS['danger']))
            self.puan_tablo.setItem(row, 2, seviye_item)
            
            self.puan_tablo.setItem(row, 3, QTableWidgetItem(p['son_guncelleme']))
        
        # Odul Tablosu
        self.odul_tablo.setRowCount(0)
        for o in oduller:
            row = self.odul_tablo.rowCount()
            self.odul_tablo.insertRow(row)
            
            self.odul_tablo.setItem(row, 0, QTableWidgetItem(str(o['id'])))
            self.odul_tablo.setItem(row, 1, QTableWidgetItem(o['ad']))
            self.odul_tablo.setItem(row, 2, QTableWidgetItem(o['aciklama'] or ""))
            self.odul_tablo.setItem(row, 3, QTableWidgetItem(f"{o['puan_maliyeti']:,}"))
            
            stok_item = QTableWidgetItem(str(o['stok']))
            if o['stok'] == 0:
                stok_item.setForeground(QColor(COLORS['danger']))
            elif o['stok'] < 10:
                stok_item.setForeground(QColor(COLORS['warning']))
            else:
                stok_item.setForeground(QColor(COLORS['success']))
            self.odul_tablo.setItem(row, 4, stok_item)
        
        # Hediye Tablosu
        self.hediye_tablo.setRowCount(0)
        for h in hediyeler[:50]:
            row = self.hediye_tablo.rowCount()
            self.hediye_tablo.insertRow(row)
            
            self.hediye_tablo.setItem(row, 0, QTableWidgetItem(str(h['id'])))
            self.hediye_tablo.setItem(row, 1, QTableWidgetItem(h['musteri_adi']))
            self.hediye_tablo.setItem(row, 2, QTableWidgetItem(h['hediye_adi']))
            self.hediye_tablo.setItem(row, 3, QTableWidgetItem(f"{h['puan_maliyeti']:,}"))
            self.hediye_tablo.setItem(row, 4, QTableWidgetItem(h['tarih']))
    
    def cikis_yap(self):
        cevap = QMessageBox.question(self, "Cikis", "Cikis yapmak istediginizden emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        if cevap == QMessageBox.Yes:
            sys.exit()

def main():
    app = QApplication(sys.argv)
    window = CRMMainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
